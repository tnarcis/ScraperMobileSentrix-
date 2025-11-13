from flask import Flask, request, jsonify, render_template, send_file, Response, stream_template
import requests, re, json, io, time, os, socket, datetime, csv, uuid, hashlib
from urllib.parse import urlparse
from dataclasses import asdict
from typing import List, Dict, Tuple, Optional, Set, Any
from openpyxl import Workbook, load_workbook
from PIL import Image
import base64
import pytz
from concurrent.futures import ThreadPoolExecutor
import threading
from database import db_manager, results_db_manager
from logger import log_job_start, log_job_complete, log_job_error

from bs4 import BeautifulSoup

# Import scraper engines (separated for maintainability)
from scraper_engine import (
    Item, CategoryInfo, build_session, scrape_url, scrape_urls_parallel,
    clean_text, parse_price_number, fmt_price, host_currency,
    scrape_category_all_pages, discover_mobilesentrix_categories,
    discover_xcell_categories, scrape_category_with_pagination,
    normalize_stock_status, extract_stock_status_from_page, PARSER
)

# Import XCellParts specialized scraper
import xcell_scraper_engine

# Import TXParts specialized scraper
import txparts_scraper_engine

app = Flask(__name__)
ADMIN_TOKEN = os.environ.get('ADMIN_TOKEN')
SUPPORTED_RESULTS_CLIENTS = {'mobilesentrix', 'xcellparts', 'txparts'}
_category_limit_env = os.getenv('MSX_CATEGORY_AUTO_LIMIT')
try:
    CATEGORY_AUTOMATIC_LIMIT = int(_category_limit_env) if _category_limit_env is not None else 150
except (TypeError, ValueError):
    CATEGORY_AUTOMATIC_LIMIT = 150
if CATEGORY_AUTOMATIC_LIMIT < 0:
    CATEGORY_AUTOMATIC_LIMIT = 0
# Soft cap to keep default MobileSentrix runs manageable without user scoping.

# ========== BACKGROUND JOB EXECUTOR ==========

# Thread pool for background jobs
executor = ThreadPoolExecutor(max_workers=4)

# In-memory job tracking
JOBS = {}
JOBS_LOCK = threading.Lock()

# Lightweight cache/session for detail stock lookups so we avoid re-fetching the
# same product page repeatedly when the listing card does not expose stock data.
DETAIL_STOCK_CACHE: Dict[str, Optional[str]] = {}
DETAIL_STOCK_CACHE_LOCK = threading.Lock()
DETAIL_STOCK_SESSION = None
DETAIL_STOCK_SESSION_LOCK = threading.Lock()

class JobStatus:
    """Track background job status"""

    def __init__(self, job_id: str, client: str, config: dict):
        self.job_id = job_id
        self.client = client
        self.config = config
        self.status = "queued"
        self.pages_done = 0
        self.items_found = 0
        self.last_error: Optional[str] = None
        self.started_at: Optional[datetime.datetime] = None
        self.completed_at: Optional[datetime.datetime] = None
        self.total_categories = 0
        self.categories_done = 0
        self.current_category: Optional[str] = None
        self.new_products = 0
        self.updated_products = 0
        self.run_id: Optional[str] = None
        self.cancel_requested = False
        self.cancel_reason: Optional[str] = None
        self.cancelled_at: Optional[datetime.datetime] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "job_id": self.job_id,
            "client": self.client,
            "status": self.status,
            "pages_done": self.pages_done,
            "items_found": self.items_found,
            "last_error": self.last_error,
            "started_at": self.started_at.isoformat() if self.started_at else None,
            "completed_at": self.completed_at.isoformat() if self.completed_at else None,
            "total_categories": self.total_categories,
            "categories_done": self.categories_done,
            "current_category": self.current_category,
            "new_products": self.new_products,
            "updated_products": self.updated_products,
            "run_id": self.run_id,
            "cancel_requested": self.cancel_requested,
            "cancel_reason": self.cancel_reason,
            "cancelled_at": self.cancelled_at.isoformat() if self.cancelled_at else None,
        }

    def request_cancel(self, reason: Optional[str] = None) -> None:
        self.cancel_requested = True
        self.cancel_reason = reason or "Scrape cancelled by user."


def slugify(value: str) -> str:
    """Simplified slugify helper for database keys."""
    if not value:
        return "general"
    normalized = clean_text(value).lower()
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized)
    return normalized.strip('-') or "general"


def derive_taxonomy(category_info: Optional[CategoryInfo], item: Item) -> Tuple[str, str, str, List[str]]:
    """Derive brand/category/model labels from breadcrumbs and discovery metadata."""
    breadcrumbs: List[str] = []
    if item.category_path:
        breadcrumbs = [clean_text(part) for part in item.category_path.split('>') if clean_text(part)]

    if not breadcrumbs and category_info and category_info.label_text:
        breadcrumbs = [clean_text(category_info.label_text)]

    brand_name = (
        clean_text(category_info.brand)
        if category_info and category_info.brand
        else (breadcrumbs[0] if breadcrumbs else "MobileSentrix")
    )

    if len(breadcrumbs) >= 3:
        category_name = breadcrumbs[1]
        model_name = breadcrumbs[-1]
    elif len(breadcrumbs) == 2:
        category_name = breadcrumbs[0]
        model_name = breadcrumbs[1]
    elif len(breadcrumbs) == 1:
        category_name = breadcrumbs[0]
        model_name = breadcrumbs[0]
    else:
        category_name = clean_text(category_info.label_text) if category_info else "General"
        model_name = category_name

    return brand_name or "MobileSentrix", category_name or "General", model_name or "General", breadcrumbs


def derive_sku(item: Item) -> str:
    """Generate a deterministic SKU identifier for result storage."""
    if item.sku and clean_text(item.sku):
        candidate = clean_text(item.sku)
    else:
        parsed = urlparse(item.url or "")
        path = (parsed.path or "").rstrip('/')
        candidate = path.split('/')[-1] or path.replace('/', '-')
        if not candidate:
            candidate = hashlib.md5((item.url or str(time.time())).encode('utf-8')).hexdigest()
    candidate = re.sub(r"[^A-Za-z0-9_-]+", "-", candidate).strip('-')
    if not candidate:
        candidate = hashlib.md5((item.url or f"sku-{time.time()}").encode('utf-8')).hexdigest()
    return candidate.upper()


def filter_categories_by_seed(categories: List[CategoryInfo], seed_url: str) -> List[CategoryInfo]:
    """Optional seed filtering so jobs can target a specific branch."""
    if not seed_url:
        return categories

    normalized = seed_url.strip()
    if not normalized:
        return categories

    normalized_lower = normalized.rstrip('/').lower()
    matches = [
        cat for cat in categories
        if normalized_lower in cat.url.lower() or cat.url.lower().startswith(normalized_lower)
    ]

    if matches:
        return matches

    label_hint = clean_text(normalized.rstrip('/').split('/')[-1].replace('-', ' ')) or "Direct Category"
    return [CategoryInfo(
        site='mobilesentrix',
        url=normalized,
        brand=None,
        label_text=label_hint,
        discovered_at=time.strftime('%Y-%m-%d %H:%M:%S')
    )]


def _get_detail_stock_session():
    """Lazily build a hardened session for product detail lookups."""
    global DETAIL_STOCK_SESSION
    with DETAIL_STOCK_SESSION_LOCK:
        if DETAIL_STOCK_SESSION is None:
            session, _ = build_session(retries=2, verify_ssl=True, use_curl=False)
            DETAIL_STOCK_SESSION = session
    return DETAIL_STOCK_SESSION


def resolve_item_stock_status(item: Item) -> Optional[str]:
    """Best-effort stock resolution; falls back to product page when needed."""
    normalized = normalize_stock_status(item.stock_status)
    if normalized:
        return normalized

    product_url = (item.url or '').strip()
    if not product_url:
        return None

    with DETAIL_STOCK_CACHE_LOCK:
        if product_url in DETAIL_STOCK_CACHE:
            return DETAIL_STOCK_CACHE[product_url]

    session = _get_detail_stock_session()
    if session is None:
        return None

    resolved: Optional[str] = None
    try:
        response = session.get(product_url, timeout=25)
        if response.ok and response.text:
            soup = BeautifulSoup(response.text, PARSER)
            detail_status = extract_stock_status_from_page(soup)
            resolved = normalize_stock_status(detail_status)
    except Exception:
        resolved = None

    with DETAIL_STOCK_CACHE_LOCK:
        DETAIL_STOCK_CACHE[product_url] = resolved

    return resolved


def upsert_mobilesentrix_items(
    category: CategoryInfo,
    items: List[Item],
    seen_skus: Set[str]
) -> Dict[str, Any]:
    """Persist items into the results database and return summary stats."""
    summary = {
        'saved': 0,
        'new': 0,
        'updated': 0,
        'brand_ids': set(),
        'category_ids': set(),
        'model_ids': set(),
        'last_brand': None,
        'last_category': category.label_text,
        'last_model': None,
        'errors': []
    }

    base_host = ""
    if category.url:
        parsed = urlparse(category.url)
        base_host = f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme and parsed.netloc else category.url

    for item in items:
        try:
            sku = derive_sku(item)
            if not sku or sku in seen_skus:
                continue
            seen_skus.add(sku)

            brand_name, category_name, model_name, breadcrumbs = derive_taxonomy(category, item)

            brand_id = results_db_manager.save_brand(
                brand_name,
                slugify(brand_name),
                base_host or category.url
            )

            category_id = results_db_manager.save_category(
                brand_id,
                category_name,
                slugify(category_name),
                category.url
            )

            model_id = results_db_manager.save_model(
                category_id,
                model_name,
                slugify(model_name),
                item.source or category.url
            )

            resolved_stock = resolve_item_stock_status(item)

            product_data = {
                'model_id': model_id,
                'sku': sku,
                'title': item.title or sku,
                'description': '',
                'price': item.price_value,
                'stock_status': resolved_stock or '',
                'availability': '',
                'condition': '',
                'product_url': item.url,
                'image_urls': [item.image_url] if item.image_url else [],
                'variant_details': {},
                'compatibility': breadcrumbs,
                'bulk_discounts': {}
            }

            product_id, is_new = results_db_manager.save_product(product_data)
            if not product_id:
                continue

            summary['saved'] += 1
            if is_new:
                summary['new'] += 1
            else:
                summary['updated'] += 1

            summary['brand_ids'].add(brand_id)
            summary['category_ids'].add(category_id)
            summary['model_ids'].add(model_id)
            summary['last_brand'] = brand_name
            summary['last_category'] = category_name
            summary['last_model'] = model_name

        except Exception as item_error:
            summary['errors'].append(str(item_error))

    return summary


def run_background_scrape(
    job_id: str,
    client: str,
    seed_url: str,
    max_pages: int,
    selected_categories: Optional[List[str]] = None
):
    """Background scraping task"""
    with JOBS_LOCK:
        if job_id not in JOBS:
            return
        job = JOBS[job_id]
        job.status = "running"
        job.started_at = datetime.datetime.now(datetime.timezone.utc)

    requested_categories = [
        url.strip()
        for url in (selected_categories or [])
        if isinstance(url, str) and url.strip()
    ]

    seed_supplied = bool(seed_url.strip()) if isinstance(seed_url, str) else False

    log_job_start(
        job_id,
        client,
        seed_url=seed_url,
        max_pages=max_pages,
        selected_categories_count=len(requested_categories)
    )

    try:
        try:
            per_category_limit = int(max_pages)
        except (TypeError, ValueError):
            per_category_limit = 0
        if per_category_limit <= 0:
            per_category_limit = None

        limit_for_config = per_category_limit if per_category_limit is not None else 'unlimited'
        with JOBS_LOCK:
            job.config['max_pages'] = limit_for_config
            job.config['selected_categories'] = list(requested_categories)

        def normalize_category_url(url: str) -> str:
            return (url or '').strip().rstrip('/').lower()

        def build_fallback_category(url: str, site: str) -> CategoryInfo:
            timestamp = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            label_source = (url or '').rstrip('/').split('/')[-1]
            label_hint = clean_text(label_source.replace('-', ' ')) or "Custom Category"
            return CategoryInfo(
                site=site,
                url=url,
                brand=None,
                label_text=label_hint,
                discovered_at=timestamp
            )

        def select_categories_by_url(
            discovered: List[CategoryInfo],
            urls: List[str],
            site: str
        ) -> List[CategoryInfo]:
            lookup = {normalize_category_url(cat.url): cat for cat in discovered}
            selected: List[CategoryInfo] = []

            for raw_url in urls:
                key = normalize_category_url(raw_url)
                if not key:
                    continue
                match = lookup.get(key)
                if match:
                    selected.append(match)
                else:
                    selected.append(build_fallback_category(raw_url, site))

            return selected

        selection_for_config = list(requested_categories)

        if client == "mobilesentrix":
            discovered_categories = discover_mobilesentrix_categories()
            if requested_categories:
                categories = select_categories_by_url(
                    discovered_categories,
                    requested_categories,
                    'mobilesentrix'
                )
                selection_for_config = [cat.url for cat in categories]
            else:
                categories = filter_categories_by_seed(
                    discovered_categories,
                    seed_url
                )

            auto_cap_total = len(categories)
            auto_cap_limit = None
            if (
                CATEGORY_AUTOMATIC_LIMIT > 0
                and not requested_categories
                and not seed_supplied
                and auto_cap_total > CATEGORY_AUTOMATIC_LIMIT
            ):
                categories = categories[:CATEGORY_AUTOMATIC_LIMIT]
                auto_cap_limit = CATEGORY_AUTOMATIC_LIMIT

            if not categories:
                raise RuntimeError("No MobileSentrix categories discovered for scraping.")

            with JOBS_LOCK:
                job.total_categories = len(categories)
                job.categories_done = 0
                job.current_category = None
                if requested_categories:
                    job.config['selected_categories'] = list(selection_for_config)
                if auto_cap_limit:
                    job.config['category_auto_limit'] = auto_cap_limit
                    job.config['category_auto_total'] = auto_cap_total

            run_id = job_id
            job.run_id = run_id

            run_config = {
                'client': client,
                'seed_url': seed_url,
                'max_pages_per_category': limit_for_config,
                'selected_categories': list(selection_for_config)
            }
            if auto_cap_limit:
                run_config['category_auto_limit'] = auto_cap_limit
                run_config['category_auto_total'] = auto_cap_total
            results_db_manager.create_scraper_run(run_id, run_config)

            seen_skus: Set[str] = set()
            brands_seen: Set[int] = set()
            categories_seen: Set[int] = set()
            models_seen: Set[int] = set()
            total_new_products = 0
            total_updated_products = 0
            total_products_saved = 0
            errors_accumulated: List[str] = []

            last_brand_name: Optional[str] = None
            last_category_name: Optional[str] = None
            last_model_name: Optional[str] = None
            cancelled = False

            for index, category in enumerate(categories, start=1):
                with JOBS_LOCK:
                    if job.cancel_requested:
                        cancelled = True
                        break
                    job.current_category = category.label_text
                    job.categories_done = index - 1

                try:
                    result = scrape_category_with_pagination(
                        category.url,
                        'mobilesentrix',
                        per_category_limit
                    )
                except Exception as scrape_error:
                    error_message = f"Failed to scrape {category.url}: {scrape_error}"
                    errors_accumulated.append(error_message)
                    with JOBS_LOCK:
                        job.last_error = error_message
                    continue

                persist_summary = upsert_mobilesentrix_items(category, result.items, seen_skus)

                if persist_summary.get('last_brand'):
                    last_brand_name = persist_summary['last_brand']
                if persist_summary.get('last_category'):
                    last_category_name = persist_summary['last_category']
                else:
                    last_category_name = category.label_text
                if persist_summary.get('last_model'):
                    last_model_name = persist_summary['last_model']

                total_products_saved += persist_summary['saved']
                total_new_products += persist_summary['new']
                total_updated_products += persist_summary['updated']
                brands_seen.update(persist_summary['brand_ids'])
                categories_seen.update(persist_summary['category_ids'])
                models_seen.update(persist_summary['model_ids'])
                errors_accumulated.extend(persist_summary['errors'])

                with JOBS_LOCK:
                    job.pages_done += result.total_pages
                    job.items_found += len(result.items)
                    job.categories_done = index
                    job.new_products += persist_summary['new']
                    job.updated_products += persist_summary['updated']
                    if persist_summary['errors']:
                        job.last_error = persist_summary['errors'][-1]

                results_db_manager.update_scraper_run(run_id, {
                    'total_products': total_products_saved,
                    'new_products': total_new_products,
                    'updated_products': total_updated_products,
                    'total_brands': len(brands_seen),
                    'total_categories': job.total_categories,
                    'total_models': len(models_seen),
                    'current_brand': last_brand_name,
                    'current_category': last_category_name or category.label_text,
                    'current_model': last_model_name,
                    'errors_count': len(errors_accumulated)
                })

            if cancelled:
                cancel_time = datetime.datetime.now(datetime.timezone.utc)
                with JOBS_LOCK:
                    job.status = "cancelled"
                    job.cancelled_at = cancel_time
                    job.completed_at = cancel_time
                    job.last_error = job.cancel_reason or job.last_error or "Scrape cancelled by user."

                results_db_manager.update_scraper_run(run_id, {
                    'status': 'stopped',
                    'completed_at': cancel_time,
                    'total_products': total_products_saved,
                    'new_products': total_new_products,
                    'updated_products': total_updated_products,
                    'total_brands': len(brands_seen),
                    'total_categories': job.total_categories,
                    'total_models': len(models_seen),
                    'current_brand': last_brand_name,
                    'current_category': last_category_name,
                    'current_model': last_model_name,
                    'errors_count': len(errors_accumulated)
                })

                log_job_error(job_id, client, job.last_error)
                return

            completion_updates = {
                'status': 'completed',
                'completed_at': datetime.datetime.now(),
                'total_products': total_products_saved,
                'new_products': total_new_products,
                'updated_products': total_updated_products,
                'total_brands': len(brands_seen),
                'total_categories': job.total_categories,
                'total_models': len(models_seen),
                'errors_count': len(errors_accumulated)
            }
            results_db_manager.update_scraper_run(run_id, completion_updates)

            if errors_accumulated:
                with JOBS_LOCK:
                    job.last_error = errors_accumulated[-1]

        elif client == "xcellparts":
            with JOBS_LOCK:
                if job.cancel_requested:
                    job.status = "cancelled"
                    job.cancelled_at = datetime.datetime.now(datetime.timezone.utc)
                    job.completed_at = job.cancelled_at
                    job.last_error = job.cancel_reason or "Scrape cancelled by user."
                    log_job_error(job_id, client, job.last_error)
                    return

            result = xcell_scraper_engine.scrape_all_discovered_categories(
                max_pages_per_category=per_category_limit,
                max_categories=5,
                allowed_urls=selection_for_config
            )

            returned_targets = result.get('target_urls') or []
            if returned_targets:
                selection_for_config = returned_targets

            with JOBS_LOCK:
                job.items_found = result.get('total_items', 0)
                job.pages_done = result.get('categories_scraped', 0)
                job.total_categories = result.get('categories_targeted', 0)
                job.config['selected_categories'] = list(selection_for_config)
                if result.get('errors'):
                    job.last_error = '; '.join(result['errors'])

        elif client == "txparts":
            with JOBS_LOCK:
                if job.cancel_requested:
                    job.status = "cancelled"
                    job.cancelled_at = datetime.datetime.now(datetime.timezone.utc)
                    job.completed_at = job.cancelled_at
                    job.last_error = job.cancel_reason or "Scrape cancelled by user."
                    log_job_error(job_id, client, job.last_error)
                    return

            categories_to_scrape = list(selection_for_config)

            with JOBS_LOCK:
                job.total_categories = len(categories_to_scrape)
                job.categories_done = 0
                job.items_found = 0
                job.pages_done = 0
                job.config['selected_categories'] = list(categories_to_scrape)

            if not categories_to_scrape:
                error_message = "TXParts scraping requires selecting one or more category URLs."
                with JOBS_LOCK:
                    job.status = "error"
                    job.completed_at = datetime.datetime.now(datetime.timezone.utc)
                    job.last_error = error_message
                log_job_error(job_id, client, error_message)
                return

            session = None
            all_items: List[Any] = []
            scrape_errors: List[str] = []

            try:
                session, _ = txparts_scraper_engine.build_session(retries=2, verify_ssl=True)
            except Exception as session_error:
                error_message = f"Failed to initialize TXParts session: {session_error}"
                with JOBS_LOCK:
                    job.status = "error"
                    job.completed_at = datetime.datetime.now(datetime.timezone.utc)
                    job.last_error = error_message
                log_job_error(job_id, client, error_message)
                return

            try:
                for index, url in enumerate(categories_to_scrape, start=1):
                    with JOBS_LOCK:
                        if job.cancel_requested:
                            job.status = "cancelled"
                            job.cancelled_at = datetime.datetime.now(datetime.timezone.utc)
                            job.completed_at = job.cancelled_at
                            job.last_error = job.cancel_reason or "Scrape cancelled by user."
                            break
                        job.current_category = url
                        job.categories_done = index - 1

                    try:
                        scraped_items = txparts_scraper_engine.scrape_url(
                            session,
                            url,
                            rules={},
                            crawl_pagination=False,
                            max_pages=per_category_limit or 0,
                            delay_ms=250,
                            logger=app.logger
                        )
                    except Exception as scrape_error:
                        message = f"Failed to scrape {url}: {scrape_error}"
                        scrape_errors.append(message)
                        with JOBS_LOCK:
                            job.last_error = message
                        continue

                    all_items.extend(scraped_items)

                    with JOBS_LOCK:
                        job.categories_done = index
                        job.pages_done = index
                        job.items_found += len(scraped_items)
                        job.current_category = url

                    if index < len(categories_to_scrape):
                        time.sleep(1)

                if job.cancel_requested:
                    log_job_error(job_id, client, job.last_error or "Scrape cancelled by user.")
                    return

                if all_items:
                    history_id = str(int(time.time() * 1000))
                    rules = {
                        'site': 'txparts',
                        'category_urls': categories_to_scrape,
                        'max_pages': per_category_limit
                    }

                    legacy_items = []
                    for item in all_items:
                        price_value = item.discounted if item.discounted else item.original
                        price_text = item.discounted_formatted or item.original_formatted
                        legacy_items.append({
                            'url': item.url,
                            'site': item.site or 'txparts.com',
                            'title': item.title,
                            'price_value': price_value,
                            'price_currency': 'USD',
                            'price_text': price_text,
                            'discounted_value': item.discounted,
                            'discounted_formatted': item.discounted_formatted,
                            'original_formatted': item.original_formatted,
                            'source': 'txparts_auto',
                            'image_url': item.image_url
                        })

                    results_db_manager.save_fetch_history(history_id, categories_to_scrape, legacy_items, rules)

                if scrape_errors:
                    with JOBS_LOCK:
                        job.last_error = scrape_errors[-1]

            finally:
                if session is not None:
                    try:
                        session.close()
                    except Exception:
                        pass

        with JOBS_LOCK:
            if not job.cancel_requested and job.status not in {"cancelled", "error"}:
                job.status = "done"
                job.completed_at = datetime.datetime.now(datetime.timezone.utc)

        if job.status == "done":
            log_job_complete(job_id, client, items_found=job.items_found, pages_done=job.pages_done)

    except Exception as e:
        with JOBS_LOCK:
            job.status = "error"
            job.last_error = str(e)
            job.completed_at = datetime.datetime.now(datetime.timezone.utc)

        if job.run_id:
            results_db_manager.update_scraper_run(job.run_id, {
                'status': 'failed',
                'completed_at': datetime.datetime.now(),
                'errors_count': 1,
                'error_log': [str(e)]
            })

        log_job_error(job_id, client, str(e))

# -------- Image Processing --------
def convert_image_format(image_data: bytes, source_format: str, target_format: str = 'JPEG', quality: int = 85) -> bytes:
    """Convert image from one format to another"""
    try:
        # Open image from bytes
        with Image.open(io.BytesIO(image_data)) as img:
            # Convert RGBA to RGB if necessary (for JPEG)
            if target_format.upper() == 'JPEG' and img.mode in ('RGBA', 'LA', 'P'):
                # Create white background
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background
            
            # Save to bytes buffer
            output_buffer = io.BytesIO()
            save_kwargs = {'format': target_format.upper()}
            if target_format.upper() == 'JPEG':
                save_kwargs['quality'] = quality
                save_kwargs['optimize'] = True
            elif target_format.upper() == 'PNG':
                save_kwargs['optimize'] = True
            
            img.save(output_buffer, **save_kwargs)
            return output_buffer.getvalue()
    except Exception as e:
        raise ValueError(f"Failed to convert image: {str(e)}")

def download_and_convert_image(image_url: str, target_format: str = 'JPEG', quality: int = 85) -> Tuple[bytes, str, str]:
    """Download image from URL and convert it to target format"""
    try:
        # Download image
        response = requests.get(image_url, timeout=30, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        })
        response.raise_for_status()
        
        # Detect source format
        source_format = 'UNKNOWN'
        content_type = response.headers.get('content-type', '').lower()
        if 'webp' in content_type:
            source_format = 'WEBP'
        elif 'jpeg' in content_type or 'jpg' in content_type:
            source_format = 'JPEG'
        elif 'png' in content_type:
            source_format = 'PNG'
        elif 'gif' in content_type:
            source_format = 'GIF'
        else:
            # Try to detect from URL extension
            url_lower = image_url.lower()
            if '.webp' in url_lower:
                source_format = 'WEBP'
            elif '.jpg' in url_lower or '.jpeg' in url_lower:
                source_format = 'JPEG'
            elif '.png' in url_lower:
                source_format = 'PNG'
            elif '.gif' in url_lower:
                source_format = 'GIF'
        
        # Convert image
        converted_data = convert_image_format(response.content, source_format, target_format, quality)
        
        return converted_data, source_format, target_format
        
    except Exception as e:
        raise ValueError(f"Failed to download/convert image from {image_url}: {str(e)}")

def convert_image_from_bytes(image_data: bytes, target_format: str = 'JPEG', quality: int = 85) -> Tuple[bytes, str, str]:
    """Convert image from bytes data to target format"""
    try:
        # Detect source format from image data
        source_format = 'UNKNOWN'
        try:
            with Image.open(io.BytesIO(image_data)) as img:
                source_format = img.format if img.format else 'UNKNOWN'
        except Exception:
            # Try to detect from magic bytes
            if image_data.startswith(b'\x89PNG'):
                source_format = 'PNG'
            elif image_data.startswith(b'\xff\xd8\xff'):
                source_format = 'JPEG'
            elif image_data.startswith(b'RIFF') and b'WEBP' in image_data[:12]:
                source_format = 'WEBP'
            elif image_data.startswith(b'GIF8'):
                source_format = 'GIF'
        
        # Convert image
        converted_data = convert_image_format(image_data, source_format, target_format, quality)
        
        return converted_data, source_format, target_format
        
    except Exception as e:
        raise ValueError(f"Failed to convert uploaded image: {str(e)}")

def get_image_info(image_data: bytes) -> Dict:
    """Get information about an image"""
    try:
        with Image.open(io.BytesIO(image_data)) as img:
            return {
                'format': img.format,
                'mode': img.mode,
                'size': img.size,
                'width': img.width,
                'height': img.height,
                'has_transparency': img.mode in ('RGBA', 'LA') or 'transparency' in img.info
            }
    except Exception as e:
        return {'error': str(e)}

# -------- Flask Routes --------
@app.get('/')
def index():
    return render_template('index.html')

@app.get('/history')
def history():
    return render_template('history.html')

@app.get('/image-converter')
def image_converter():
    return render_template('image_converter.html')

@app.get('/instructions')
def instructions():
    return render_template('instructions.html')

@app.get('/api/history')
def api_history():
    """Return history data from database"""
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit
        
        histories = db_manager.get_history_list(limit=limit, offset=offset)
        return jsonify({
            'histories': histories,
            'page': page,
            'limit': limit,
            'total': len(histories)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.get('/api/history/<history_id>')
def api_history_detail(history_id):
    """Return specific history entry from database"""
    try:
        history = db_manager.get_history_detail(history_id)
        if not history:
            return jsonify({'error': 'History entry not found'}), 404
        return jsonify(history)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/history/<history_id>/export/xlsx')
def api_history_export(history_id):
    """Export a specific history entry to XLSX"""
    try:
        history = db_manager.get_history_detail(history_id)
        if not history:
            return jsonify({'error': 'History entry not found'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Session"

        urls = history.get('urls', [])
        rules = history.get('rules', {}) or {}
        items = history.get('items', [])
        timestamp = history.get('timestamp')

        tz_label = "Pakistan Standard Time (UTC+05:00)"
        ts_display = timestamp
        if timestamp:
            try:
                # Parse timestamp and convert to Pakistan time
                ts_str = str(timestamp)
                if '+' in ts_str or 'Z' in ts_str:
                    # Has timezone info
                    ts_obj = datetime.datetime.fromisoformat(ts_str.replace('Z', '+00:00'))
                    if ts_obj.tzinfo is None:
                        ts_obj = pytz.UTC.localize(ts_obj)
                    # Convert to Pakistan timezone
                    pakistan_tz = pytz.timezone('Asia/Karachi')
                    ts_pakistan = ts_obj.astimezone(pakistan_tz)
                else:
                    # Assume it's already in Pakistan time
                    ts_obj = datetime.datetime.fromisoformat(ts_str)
                    pakistan_tz = pytz.timezone('Asia/Karachi')
                    ts_pakistan = pakistan_tz.localize(ts_obj)
                
                ts_display = ts_pakistan.strftime('%d %b %Y %I:%M %p PKT')
            except Exception as e:
                print(f"Error formatting timestamp in export: {e}")
                ts_display = str(timestamp)

        summary_rows = [
            ["Session ID", history.get('id', '')],
            ["Timestamp", ts_display],
            ["Timezone", tz_label],
            ["URLs Crawled", len(urls)],
            ["Items Captured", history.get('items_count', 0)],
            ["Discount %", rules.get('percent_off', 0)],
            ["Absolute Off", rules.get('absolute_off', 0)],
        ]

        for row in summary_rows:
            ws.append(row)

        if urls:
            ws.append([])
            ws.append(["Target URLs"])
            for url in urls:
                ws.append([url])

        ws.append([])
        headers = [
            "Title","Original Price","Discounted Price","Price Text","Price Value",
            "URL","Source","Site","Currency","Image URL"
        ]
        ws.append(headers)

        for item in items:
            ws.append([
                item.get('title',''),
                item.get('original_formatted',''),
                item.get('discounted_formatted',''),
                item.get('price_text',''),
                item.get('price_value',''),
                item.get('url',''),
                item.get('source',''),
                item.get('site',''),
                item.get('price_currency',''),
                item.get('image_url','')
            ])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"history_{history_id}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.delete('/api/history/<history_id>')
def api_delete_history(history_id):
    """Delete history entry from database"""
    try:
        success = db_manager.delete_history(history_id)
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'History entry not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.get('/api/statistics')
def api_statistics():
    """Get database statistics"""
    try:
        stats = db_manager.get_statistics()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/search')
def api_search():
    """Search items in database"""
    try:
        data = request.get_json(silent=True) or {}
        query = data.get('query', '').strip()
        limit = int(data.get('limit', 100))
        
        if not query:
            return jsonify({'error': 'Search query is required'}), 400
        
        items = db_manager.search_items(query, limit)
        return jsonify({
            'query': query,
            'results': items,
            'count': len(items)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/cleanup')
def api_cleanup():
    """Cleanup old database entries"""
    try:
        data = request.get_json(silent=True) or {}
        days = int(data.get('days', 90))
        
        deleted_count = db_manager.cleanup_old_entries(days)
        return jsonify({
            'success': True,
            'deleted_entries': deleted_count,
            'days': days
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/convert-image')
def api_convert_image():
    """Convert a single image from URL"""
    data = request.get_json(silent=True) or {}
    image_url = data.get('url', '').strip()
    target_format = data.get('format', 'JPEG').upper()
    quality = int(data.get('quality', 85))
    
    if not image_url:
        return jsonify({'error': 'Image URL is required'}), 400
    
    if target_format not in ['JPEG', 'PNG', 'WEBP']:
        return jsonify({'error': 'Unsupported target format'}), 400
    
    try:
        converted_data, source_format, target_format_used = download_and_convert_image(
            image_url, target_format, quality
        )
        
        # Get image info
        image_info = get_image_info(converted_data)
        
        # Convert to base64 for response
        base64_data = base64.b64encode(converted_data).decode('utf-8')
        
        return jsonify({
            'success': True,
            'original_url': image_url,
            'source_format': source_format,
            'target_format': target_format_used,
            'image_info': image_info,
            'base64_data': base64_data,
            'data_url': f"data:image/{target_format_used.lower()};base64,{base64_data}",
            'file_size': len(converted_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/convert-images-batch')
def api_convert_images_batch():
    """Convert multiple images from URLs"""
    data = request.get_json(silent=True) or {}
    image_urls = data.get('urls', [])
    target_format = data.get('format', 'JPEG').upper()
    quality = int(data.get('quality', 85))
    
    if not image_urls:
        return jsonify({'error': 'Image URLs are required'}), 400
    
    if target_format not in ['JPEG', 'PNG', 'WEBP']:
        return jsonify({'error': 'Unsupported target format'}), 400
    
    results = []
    
    for i, image_url in enumerate(image_urls[:20]):  # Limit to 20 images
        try:
            converted_data, source_format, target_format_used = download_and_convert_image(
                image_url, target_format, quality
            )
            
            image_info = get_image_info(converted_data)
            base64_data = base64.b64encode(converted_data).decode('utf-8')
            
            results.append({
                'success': True,
                'index': i,
                'original_url': image_url,
                'source_format': source_format,
                'target_format': target_format_used,
                'image_info': image_info,
                'base64_data': base64_data,
                'data_url': f"data:image/{target_format_used.lower()};base64,{base64_data}",
                'file_size': len(converted_data)
            })
            
        except Exception as e:
            results.append({
                'success': False,
                'index': i,
                'original_url': image_url,
                'error': str(e)
            })
    
    return jsonify({
        'results': results,
        'total_processed': len(results),
        'successful': len([r for r in results if r['success']]),
        'failed': len([r for r in results if not r['success']])
    })

@app.post('/api/convert-files-batch')
def api_convert_files_batch():
    """Convert multiple uploaded image files"""
    if 'files' not in request.files:
        return jsonify({'error': 'No files uploaded'}), 400
    
    files = request.files.getlist('files')
    target_format = request.form.get('format', 'JPEG').upper()
    quality = int(request.form.get('quality', 85))
    
    if not files:
        return jsonify({'error': 'No files uploaded'}), 400
    
    if target_format not in ['JPEG', 'PNG', 'WEBP']:
        return jsonify({'error': 'Unsupported target format'}), 400
    
    results = []
    
    for i, file in enumerate(files[:20]):  # Limit to 20 files
        try:
            if file.filename == '':
                results.append({
                    'success': False,
                    'index': i,
                    'filename': 'unnamed',
                    'error': 'Empty filename'
                })
                continue
                
            # Read file data
            file_data = file.read()
            
            if len(file_data) == 0:
                results.append({
                    'success': False,
                    'index': i,
                    'filename': file.filename,
                    'error': 'Empty file'
                })
                continue
            
            # Convert the uploaded file
            converted_data, source_format, target_format_used = convert_image_from_bytes(
                file_data, target_format, quality
            )
            
            image_info = get_image_info(converted_data)
            base64_data = base64.b64encode(converted_data).decode('utf-8')
            
            # Clean filename for download
            filename_base = file.filename.rsplit('.', 1)[0] if '.' in file.filename else file.filename
            
            results.append({
                'success': True,
                'index': i,
                'filename': filename_base,
                'original_filename': file.filename,
                'source_format': source_format,
                'target_format': target_format_used,
                'image_info': image_info,
                'base64_data': base64_data,
                'data_url': f"data:image/{target_format_used.lower()};base64,{base64_data}",
                'file_size': len(converted_data)
            })
            
        except Exception as e:
            results.append({
                'success': False,
                'index': i,
                'filename': file.filename if hasattr(file, 'filename') else f'file_{i}',
                'error': str(e)
            })
    
    return jsonify({
        'results': results,
        'total_processed': len(results),
        'successful': len([r for r in results if r['success']]),
        'failed': len([r for r in results if not r['success']])
    })

@app.get('/api/download-converted/<path:filename>')
def download_converted_image(filename):
    """Download converted image (would need to implement temporary storage)"""
    # This would require implementing temporary file storage
    # For now, return a simple response
    return jsonify({'error': 'Direct download not implemented yet'}), 501

@app.post('/api/scrape')
def api_scrape():
    data = request.get_json(silent=True) or {}
    urls_raw = data.get('urls') or ''
    crawl_pagination = bool(data.get('crawl_pagination', True))
    max_pages = int(data.get('max_pages') or 10)  # Reduced default from 20 to 10
    max_pages = 1 if max_pages < 1 else 20 if max_pages > 20 else max_pages  # Cap at 20 instead of 100

    delay_ms = int(data.get('delay_ms') or 50)
    retries = int(data.get('retries') or 1)
    verify_ssl = bool(data.get('verify_ssl', True))
    use_curl = bool(data.get('use_curl', False))
    use_parallel = bool(data.get('use_parallel', True))

    rules = {
        "percent_off": float(data.get('percent_off') or 0.0),
        "absolute_off": float(data.get('absolute_off') or 0.0),
    }

    urls = [u.strip() for u in (urls_raw.splitlines() if isinstance(urls_raw, str) else urls_raw) if u.strip()]
    seen_u = set(); urls = [u for u in urls if not (u in seen_u or seen_u.add(u))]

    items: List[Item] = []
    engine_used = {}  # Track which engine was used for each URL

    # Pre-create sessions per engine so we can reuse connections/cookies across URLs
    standard_session = None
    xcell_session = None
    txparts_session = None
    using_curl_flag = False
    
    # Intelligent domain routing
    def get_scraper_for_url(url: str):
        """Determine which scraper engine to use based on domain"""
        from urllib.parse import urlparse
        domain = urlparse(url).netloc.lower()
        
        # XCellParts.com uses specialized scraper
        if 'xcellparts.com' in domain:
            return 'xcell', xcell_scraper_engine
        # TXParts.com uses specialized scraper
        elif 'txparts.com' in domain:
            return 'txparts', txparts_scraper_engine
        # MobileSentrix uses standard scraper
        elif 'mobilesentrix' in domain:
            return 'standard', None
        else:
            # Default to standard scraper for unknown domains
            return 'standard', None
    
    if use_parallel and len(urls) > 1:
        # For parallel processing, group URLs by engine type
        xcell_urls = []
        txparts_urls = []
        standard_urls = []
        
        for url in urls:
            engine_type, _ = get_scraper_for_url(url)
            if engine_type == 'xcell':
                xcell_urls.append(url)
                engine_used[url] = 'xcell_scraper_engine'
            elif engine_type == 'txparts':
                txparts_urls.append(url)
                engine_used[url] = 'txparts_scraper_engine'
            else:
                standard_urls.append(url)
                engine_used[url] = 'scraper_engine'
        
        # Process XCellParts URLs with xcell scraper
        if xcell_urls:
            app.logger.info(f"[engine] Using XCellParts scraper for {len(xcell_urls)} URLs")
            if xcell_session is None:
                xcell_session, _ = xcell_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
            for url in xcell_urls:
                items.extend(xcell_scraper_engine.scrape_url(
                    xcell_session, url, rules, crawl_pagination, max_pages, delay_ms, app.logger
                ))
        
        # Process TXParts URLs with txparts scraper
        if txparts_urls:
            app.logger.info(f"[engine] Using TXParts scraper for {len(txparts_urls)} URLs")
            if txparts_session is None:
                txparts_session, _ = txparts_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
            for url in txparts_urls:
                items.extend(txparts_scraper_engine.scrape_url(
                    txparts_session, url, rules, crawl_pagination, max_pages, delay_ms, app.logger
                ))
        
        # Process standard URLs with main scraper (parallel)
        if standard_urls:
            app.logger.info(f"[engine] Using standard scraper for {len(standard_urls)} URLs")
            if standard_session is None:
                standard_session, using_curl = build_session(retries=retries, verify_ssl=verify_ssl, use_curl=use_curl)
            headers_snapshot = dict(standard_session.headers)
            cookies_snapshot = standard_session.cookies.copy()

            def standard_session_factory():
                sess, _ = build_session(retries=retries, verify_ssl=verify_ssl, use_curl=use_curl)
                sess.headers.update(headers_snapshot)
                try:
                    sess.cookies.update(cookies_snapshot)
                except Exception:
                    pass
                return sess

            max_workers = min(3, len(standard_urls))
            items.extend(scrape_urls_parallel(
                standard_urls,
                rules,
                crawl_pagination,
                max_pages,
                delay_ms,
                retries,
                verify_ssl,
                use_curl,
                max_workers,
                app.logger,
                session_factory=standard_session_factory
            ))
        
        using_curl = use_curl
    else:
        # Sequential processing
        for u in urls:
            try:
                if use_parallel and len(urls) > 1:
                    # For parallel processing, group URLs by engine type
                    xcell_urls = []
                    txparts_urls = []
                    standard_urls = []

                    for url in urls:
                        engine_type, _ = get_scraper_for_url(url)
                        if engine_type == 'xcell':
                            xcell_urls.append(url)
                            engine_used[url] = 'xcell_scraper_engine'
                        elif engine_type == 'txparts':
                            txparts_urls.append(url)
                            engine_used[url] = 'txparts_scraper_engine'
                        else:
                            standard_urls.append(url)
                            engine_used[url] = 'scraper_engine'

                    # Process XCellParts URLs with xcell scraper
                    if xcell_urls:
                        app.logger.info(f"[engine] Using XCellParts scraper for {len(xcell_urls)} URLs")
                        if xcell_session is None:
                            xcell_session, _ = xcell_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
                        for url in xcell_urls:
                            items.extend(xcell_scraper_engine.scrape_url(
                                xcell_session, url, rules, crawl_pagination, max_pages, delay_ms, app.logger
                            ))

                    # Process TXParts URLs with txparts scraper
                    if txparts_urls:
                        app.logger.info(f"[engine] Using TXParts scraper for {len(txparts_urls)} URLs")
                        if txparts_session is None:
                            txparts_session, _ = txparts_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
                        for url in txparts_urls:
                            items.extend(txparts_scraper_engine.scrape_url(
                                txparts_session, url, rules, crawl_pagination, max_pages, delay_ms, app.logger
                            ))

                    # Process standard URLs with main scraper (parallel)
                    if standard_urls:
                        app.logger.info(f"[engine] Using standard scraper for {len(standard_urls)} URLs")
                        if standard_session is None:
                            standard_session, session_is_curl = build_session(retries=retries, verify_ssl=verify_ssl, use_curl=use_curl)
                            using_curl_flag = using_curl_flag or session_is_curl
                        headers_snapshot = dict(standard_session.headers)
                        cookies_snapshot = standard_session.cookies.copy()

                        def standard_session_factory():
                            nonlocal using_curl_flag
                            sess, session_is_curl = build_session(retries=retries, verify_ssl=verify_ssl, use_curl=use_curl)
                            sess.headers.update(headers_snapshot)
                            try:
                                sess.cookies.update(cookies_snapshot)
                            except Exception:
                                pass
                            using_curl_flag = using_curl_flag or session_is_curl
                            return sess

                        max_workers = min(3, len(standard_urls))
                        items.extend(scrape_urls_parallel(
                            standard_urls,
                            rules,
                            crawl_pagination,
                            max_pages,
                            delay_ms,
                            retries,
                            verify_ssl,
                            use_curl,
                            max_workers,
                            app.logger,
                            session_factory=standard_session_factory
                        ))

                else:
                    # Sequential processing
                    for u in urls:
                        engine_type, _ = get_scraper_for_url(u)

                        if engine_type == 'xcell':
                            # Use XCellParts scraper
                            app.logger.info(f"[engine] Using XCellParts scraper for: {u}")
                            engine_used[u] = 'xcell_scraper_engine'
                            if xcell_session is None:
                                xcell_session, _ = xcell_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
                            items.extend(xcell_scraper_engine.scrape_url(
                                xcell_session, u, rules, crawl_pagination, max_pages, delay_ms, app.logger
                            ))
                        elif engine_type == 'txparts':
                            # Use TXParts scraper
                            app.logger.info(f"[engine] Using TXParts scraper for: {u}")
                            engine_used[u] = 'txparts_scraper_engine'
                            if txparts_session is None:
                                txparts_session, _ = txparts_scraper_engine.build_session(retries=retries, verify_ssl=verify_ssl)
                            items.extend(txparts_scraper_engine.scrape_url(
                                txparts_session, u, rules, crawl_pagination, max_pages, delay_ms, app.logger
                            ))
                        else:
                            # Use standard scraper
                            app.logger.info(f"[engine] Using standard scraper for: {u}")
                            engine_used[u] = 'scraper_engine'
                            if standard_session is None:
                                standard_session, session_is_curl = build_session(retries=retries, verify_ssl=verify_ssl, use_curl=use_curl)
                                using_curl_flag = using_curl_flag or session_is_curl
                            items.extend(scrape_url(standard_session, u, rules, crawl_pagination, max_pages, delay_ms, app.logger))

                # Store in database instead of memory
                history_id = str(int(time.time() * 1000))  # timestamp-based ID

                # Save to database in background thread to not block response
                def save_to_db():
                    try:
                        success = db_manager.save_fetch_history(history_id, urls, items, rules)
                        if not success:
                            app.logger.error("Failed to save fetch history to database")
                    except Exception as e:
                        app.logger.error(f"Database error: {e}")

                # Start background save
                threading.Thread(target=save_to_db, daemon=True).start()

                return jsonify({
                    "rules": rules,
                    "count": len(items),
                    "using_curl": using_curl_flag,
                    "using_parallel": use_parallel and len(urls) > 1,
                    "engines_used": engine_used,  # Show which engine was used for each URL
                    "items": [asdict(i) for i in items],
                    "history_id": history_id
                })

            except Exception as exc:
                app.logger.exception("Extractor scrape failed")
                return jsonify({'error': str(exc)}), 500

            finally:
                for sess in (standard_session, xcell_session, txparts_session):
                    if sess is not None:
                        try:
                            sess.close()
                        except Exception:
                            pass
        

@app.post('/api/comparison/upload')
def api_comparison_upload():
    """Parse an uploaded comparison sheet (CSV/XLSX) and return normalized rows."""
    uploaded = request.files.get('file')
    if uploaded is None or uploaded.filename.strip() == '':
        return jsonify({'status': 'error', 'error': 'No file uploaded.'}), 400

    raw = uploaded.read()
    if not raw:
        return jsonify({'status': 'error', 'error': 'The uploaded file is empty.'}), 400

    filename = uploaded.filename
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    title_fields = (
        'title', 'product', 'product title', 'product name', 'name', 'item', 'item name'
    )
    price_fields = (
        'price', 'cost', 'amount', 'usd', 'usd price', 'price usd', 'current price'
    )
    site_fields = (
        'site', 'source', 'vendor', 'store', 'website'
    )
    url_fields = (
        'url', 'link', 'product url', 'product link', 'href'
    )

    price_sanitize_re = re.compile(r'[^0-9.,-]+')

    def normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
        normalized: Dict[str, Any] = {}
        for key, value in row.items():
            if key is None:
                continue
            header = str(key).strip().lower()
            if not header:
                continue
            normalized[header] = value
        return normalized

    def parse_price(value: Any) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            val = float(value)
            if val != val or val in (float('inf'), float('-inf')):
                return None
            return val

        text = str(value).strip()
        if not text:
            return None

        cleaned = price_sanitize_re.sub('', text.replace(' ', ''))
        if not cleaned:
            return None

        if '.' in cleaned and ',' in cleaned:
            cleaned = cleaned.replace(',', '')
        elif ',' in cleaned:
            parts = cleaned.split(',')
            if len(parts[-1]) <= 2:
                cleaned = '.'.join(parts)
            else:
                cleaned = ''.join(parts)
        else:
            cleaned = cleaned.replace(',', '')

        try:
            return float(cleaned)
        except ValueError:
            if cleaned.count('.') > 1:
                left, _, right = cleaned.rpartition('.')
                cleaned = left.replace('.', '') + ('.' + right if right else '')
                try:
                    return float(cleaned)
                except ValueError:
                    return None
            return None

    def extract_row(normalized_row: Dict[str, Any]) -> Optional[Dict[str, object]]:
        title_value = ''
        for field in title_fields:
            value = normalized_row.get(field)
            if value in (None, ''):
                continue
            text = str(value).strip()
            if text:
                title_value = text
                break
        if not title_value:
            return None

        price_value: Optional[float] = None
        for field in price_fields:
            price_value = parse_price(normalized_row.get(field))
            if price_value is not None:
                break
        if price_value is None:
            return None

        site_value = ''
        for field in site_fields:
            value = normalized_row.get(field)
            if value in (None, ''):
                continue
            text = str(value).strip()
            if text:
                site_value = text
                break

        url_value = ''
        for field in url_fields:
            value = normalized_row.get(field)
            if value in (None, ''):
                continue
            text = str(value).strip()
            if text:
                url_value = text
                break

        return {
            'title': title_value,
            'price': float(round(price_value, 4)),
            'site': site_value,
            'url': url_value
        }

    extracted: List[Dict[str, object]] = []
    skipped = 0

    try:
        if ext in {'csv', 'txt'}:
            text = raw.decode('utf-8-sig', errors='ignore')
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise ValueError('No headers found in CSV file.')
            for row in reader:
                normalized = normalize_row(row)
                if not normalized:
                    continue
                result = extract_row(normalized)
                if result is None:
                    skipped += 1
                    continue
                extracted.append(result)
        else:
            workbook = load_workbook(io.BytesIO(raw), data_only=True)
            sheet = workbook.active
            rows_iter = list(sheet.iter_rows(values_only=True))
            if not rows_iter:
                raise ValueError('Spreadsheet is empty.')

            headers_raw = rows_iter[0]
            headers = [str(h or '').strip().lower() for h in headers_raw]
            if not any(headers):
                raise ValueError('Header row is missing in the spreadsheet.')

            for row_values in rows_iter[1:]:
                row_dict: Dict[str, Any] = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = row_values[idx] if idx < len(row_values) else None
                    row_dict[header] = value
                if not row_dict:
                    continue
                result = extract_row(row_dict)
                if result is None:
                    skipped += 1
                    continue
                extracted.append(result)
    except ValueError as ve:
        return jsonify({'status': 'error', 'error': str(ve)}), 400
    except Exception as exc:
        return jsonify({'status': 'error', 'error': f'Failed to process file: {exc}'}), 400

    if not extracted:
        return jsonify({'status': 'error', 'error': 'No valid rows found. Ensure the file includes both title and price columns.'}), 400

    message = f"Loaded {len(extracted)} comparison rows"
    if skipped:
        message += f" (skipped {skipped} rows without title or price)"

    return jsonify({'status': 'success', 'message': message, 'rows': extracted})

# -------- Main --------
def find_free_port(start=5000, end=5050):
    for p in range(start, end + 1):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                s.bind(('0.0.0.0', p))
                return p
            except OSError:
                continue
    return 0

# ========== RESULTS DASHBOARD API ROUTES ==========

@app.get('/results')
def results_dashboard():
    """Results dashboard page"""
    return render_template('results.html')

@app.get('/api/health')
def api_health():
    """Health check endpoint"""
    try:
        # Test database connection
        db_connected = True
        db_last_write = None
        
        try:
            conn = db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT MAX(created_at) FROM items')
            result = cursor.fetchone()
            db_last_write = result[0] if result and result[0] else None
        except Exception:
            db_connected = False
        
        return jsonify({
            "ok": True,
            "db_connected": db_connected,
            "db_last_write_iso": db_last_write,
            "version": "v8"
        })
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e),
            "version": "v8"
        }), 500

@app.post('/api/scrape/start')
def api_scrape_start():
    """Start background scraping job"""
    try:
        data = request.get_json() or {}
        client = data.get('client', 'mobilesentrix')
        seed_url = data.get('seed_url', '')
        max_pages_raw = data.get('max_pages', 0)
        try:
            max_pages = int(max_pages_raw)
        except (TypeError, ValueError):
            max_pages = 0
        if max_pages < 0:
            max_pages = 0

        raw_categories = data.get('categories') or []
        if raw_categories is None:
            raw_categories = []
        if not isinstance(raw_categories, list):
            return jsonify({'error': 'categories must be an array of URLs'}), 400

        selected_categories = [
            url.strip()
            for url in raw_categories
            if isinstance(url, str) and url.strip()
        ]
        
        if client not in SUPPORTED_RESULTS_CLIENTS:
            return jsonify({'error': f'Unsupported client "{client}" for scraping'}), 400
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        
        # Create job status
        job = JobStatus(job_id, client, {
            'seed_url': seed_url,
            'max_pages': max_pages,
            'selected_categories': list(selected_categories)
        })
        
        with JOBS_LOCK:
            JOBS[job_id] = job

        # Submit to executor
        executor.submit(
            run_background_scrape,
            job_id,
            client,
            seed_url,
            max_pages,
            selected_categories
        )
        
        return jsonify({
            'job_id': job_id,
            'status': 'queued',
            'client': client,
            'selected_categories': selected_categories
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.get('/api/scrape/categories')
def api_scrape_categories():
    """Return discoverable category links for the requested client."""
    client = request.args.get('client', 'mobilesentrix')
    if client not in SUPPORTED_RESULTS_CLIENTS:
        return jsonify({'error': f'Unsupported client "{client}" for category discovery'}), 400

    try:
        if client == 'mobilesentrix':
            categories = discover_mobilesentrix_categories()
        elif client == 'xcellparts':
            categories = discover_xcell_categories()
        else:  # txparts lacks automated discovery for now
            return jsonify({
                'client': client,
                'count': 0,
                'categories': [],
                'warning': 'Category discovery for TXParts is not yet automated. Provide specific URLs manually.'
            })

        payload = [asdict(category) for category in categories]
        return jsonify({
            'client': client,
            'count': len(payload),
            'categories': payload
        })
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500

@app.get('/api/scrape/status')
def api_scrape_status():
    """Get scraping job status"""
    job_id = request.args.get('job_id')
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400
    
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
        
        return jsonify({
            "status": job.status,
            "pages_done": job.pages_done,
            "items_found": job.items_found,
            "categories_done": job.categories_done,
            "total_categories": job.total_categories,
            "current_category": job.current_category,
            "new_products": job.new_products,
            "updated_products": job.updated_products,
            "run_id": job.run_id,
            "started_at": job.started_at.isoformat() if job.started_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
            "last_error": job.last_error,
            "cancel_requested": job.cancel_requested,
            "cancel_reason": job.cancel_reason,
            "cancelled_at": job.cancelled_at.isoformat() if job.cancelled_at else None,
            "config": dict(job.config)
        })


@app.post('/api/scrape/stop')
def api_scrape_stop():
    """Request cancellation of an active scraping job"""
    data = request.get_json(silent=True) or {}
    job_id = data.get('job_id')
    if not job_id:
        return jsonify({'error': 'job_id required'}), 400

    reason = data.get('reason')

    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404

        if job.status in {"done", "completed", "error", "cancelled"}:
            return jsonify({
                'status': job.status,
                'message': 'Job already finished.'
            })

        job.request_cancel(reason)
        job.last_error = job.cancel_reason
        cancel_message = job.cancel_reason

    return jsonify({
        'status': 'cancelling',
        'message': cancel_message
    })

@app.get('/api/results/summary')
def api_results_summary():
    """Get results dashboard summary data"""
    try:
        client = request.args.get('client', 'mobilesentrix')
        if client not in SUPPORTED_RESULTS_CLIENTS:
            client = 'mobilesentrix'
        
        # Get basic data using the results database to keep dashboards isolated
        clients = results_db_manager.get_clients()
        totals = results_db_manager.get_totals(client)
        categories = results_db_manager.get_category_completion(client)
        last_run = results_db_manager.get_last_run(client)
        next_run_eta = results_db_manager.get_next_run_eta_minutes(client)
        
        # Check if any job is currently running
        current_job = None
        with JOBS_LOCK:
            for job_id, job in JOBS.items():
                if job.client == client and job.status in ['queued', 'running']:
                    current_job = {
                        "status": job.status,
                        "pages_done": job.pages_done,
                        "items_found": job.items_found,
                        "categories_done": job.categories_done,
                        "total_categories": job.total_categories,
                        "current_category": job.current_category,
                        "new_products": job.new_products,
                        "updated_products": job.updated_products,
                        "job_id": job_id,
                        "client": job.client,
                        "run_id": job.run_id,
                        "started_at": job.started_at.isoformat() if job.started_at else None,
                        "last_error": job.last_error,
                        "config": dict(job.config)
                    }
                    break
        
        if not current_job:
            current_job = {
                "status": "idle",
                "pages_done": 0,
                "items_found": 0,
                "categories_done": 0,
                "total_categories": 0,
                "current_category": None,
                "new_products": 0,
                "updated_products": 0,
                "job_id": None,
                "client": client,
                "run_id": None,
                "started_at": None,
                "last_error": None,
                "config": {
                    'selected_categories': [],
                    'max_pages': None
                }
            }
        else:
            if current_job.get("total_categories"):
                total = current_job["total_categories"]
                completed = current_job.get("categories_done", 0)
                pct = round((completed / max(total, 1)) * 100, 2)
                categories = {
                    "total": total,
                    "completed": completed,
                    "completion_pct": pct
                }
        
        return jsonify({
            "clients": clients,
            "active_client": client,
            "totals": totals,
            "categories": categories,
            "runs": {
                "last_run_at": last_run,
                "next_run_eta_minutes": next_run_eta
            },
            "job": current_job
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.get('/api/results/recent')
def api_results_recent():
    """Get recent changes with filters"""
    try:
        client = request.args.get('client', 'mobilesentrix')
        if client not in SUPPORTED_RESULTS_CLIENTS:
            client = 'mobilesentrix'
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        change_types = request.args.getlist('change_types') or ['price', 'stock', 'description']
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        search_query = request.args.get('q', '').strip()
        
        changes_result = results_db_manager.get_recent_changes(
            client=client,
            limit=limit,
            offset=offset,
            change_types=change_types,
            from_date=from_date,
            to_date=to_date,
            search_query=search_query if search_query else None,
            with_total=True
        )
        
        return jsonify({
            "changes": changes_result.get("items", []),
            "total": changes_result.get("total", 0),
            "limit": limit,
            "offset": offset
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.post('/api/admin/purge')
def api_admin_purge():
    """Administrative purge for old results data."""
    data = request.get_json(silent=True) or {}

    if ADMIN_TOKEN:
        token = data.get('token')
        if token != ADMIN_TOKEN:
            return jsonify({'error': 'Unauthorized'}), 403

    older_than_days = data.get('older_than_days')
    if older_than_days is not None and older_than_days != '':
        try:
            older_than_days = int(older_than_days)
            if older_than_days < 0:
                return jsonify({'error': 'older_than_days must be non-negative'}), 400
        except (TypeError, ValueError):
            return jsonify({'error': 'older_than_days must be an integer'}), 400
    else:
        older_than_days = None

    include_products = bool(data.get('include_products', False))
    delete_all = bool(data.get('delete_all', False))

    try:
        summary = results_db_manager.purge_results_data(
            older_than_days=older_than_days,
            include_products=include_products,
            delete_all=delete_all
        )

        return jsonify({
            'status': 'ok',
            'deleted': summary,
            'older_than_days': older_than_days,
            'include_products': include_products,
            'delete_all': delete_all
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.post('/api/results/export/xlsx')
def api_results_export_xlsx():
    """Export changes to XLSX"""
    try:
        data = request.get_json() or {}
        client = data.get('client', 'mobilesentrix')
        if client not in SUPPORTED_RESULTS_CLIENTS:
            client = 'mobilesentrix'
        filters = data.get('filters', {})
        
        # Generate XLSX export from the results database
        xlsx_buffer = results_db_manager.export_changes_to_xlsx(client, filters)
        
        # Return file
        return send_file(
            xlsx_buffer,
            as_attachment=True,
            download_name=f"{client}_changes_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port_env = os.environ.get('PORT')
    port = int(port_env) if port_env else find_free_port()
    if not port:
        raise SystemExit("No free port in 5000–5050. Set PORT env var to a free port.")

    is_production = os.getenv('FLY_APP_NAME') is not None
    app.run(host='0.0.0.0', port=port, debug=not is_production)
