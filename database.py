"""
Database module for MobileSentrix Extractor
Handles persistent storage of scraping history and items
"""

import sqlite3
import json
import datetime
import io
import re
from typing import List, Optional, Dict, Any, Tuple
from dataclasses import asdict
import threading
import pytz
import os

# Pakistan timezone for consistent handling
PAKISTAN_TZ = pytz.timezone('Asia/Karachi')

def get_pakistan_time(dt=None):
    """Get current time in Pakistan timezone or convert a datetime to Pakistan timezone"""
    if dt is None:
        return datetime.datetime.now(PAKISTAN_TZ)
    if dt.tzinfo is None:
        # Assume UTC if no timezone info
        dt = pytz.UTC.localize(dt)
    return dt.astimezone(PAKISTAN_TZ)

def utc_to_pakistan(utc_dt):
    """Convert UTC datetime to Pakistan timezone"""
    if utc_dt.tzinfo is None:
        utc_dt = pytz.UTC.localize(utc_dt)
    return utc_dt.astimezone(PAKISTAN_TZ)

class DatabaseManager:
    def __init__(
        self,
        db_path: str = None,
        create_legacy_tables: bool = True,
        create_results_tables: bool = True
    ):
        self._local = threading.local()
        self.create_legacy_tables = create_legacy_tables
        self.create_results_tables = create_results_tables

        configured_path = db_path or os.environ.get("DATABASE_PATH")
        if configured_path:
            self.db_path = configured_path
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            self.db_path = os.path.join(base_dir, "mobilesentrix.db")

        db_dir = os.path.dirname(self.db_path)
        if db_dir and not os.path.exists(db_dir):
            os.makedirs(db_dir, exist_ok=True)

        # In-memory caches to reduce redundant lookups during bulk imports
        self._cache_lock = threading.Lock()
        self._brand_cache: Dict[str, int] = {}
        self._category_cache: Dict[Tuple[int, str], int] = {}
        self._model_cache: Dict[Tuple[int, str], int] = {}
        
        self.init_database()
    
    def get_connection(self):
        """Get thread-local database connection"""
        if not hasattr(self._local, 'connection'):
            conn = sqlite3.connect(self.db_path, check_same_thread=False)
            conn.row_factory = sqlite3.Row  # Enable dict-like access
            conn.execute('PRAGMA foreign_keys = ON')
            self._local.connection = conn
        return self._local.connection
    
    def init_database(self):
        """Initialize database tables"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        if self.create_legacy_tables:
            self._init_legacy_tables(cursor)
        
        if self.create_results_tables:
            self._init_results_tables(cursor)
        
        conn.commit()
    
    def _init_legacy_tables(self, cursor):
        """Create legacy extractor tables"""
        # Create fetch_history table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS fetch_history (
                id TEXT PRIMARY KEY,
                timestamp DATETIME NOT NULL,
                urls TEXT NOT NULL,  -- JSON array of URLs
                items_count INTEGER NOT NULL,
                rules TEXT NOT NULL,  -- JSON object with scraping rules
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create items table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                history_id TEXT NOT NULL,
                url TEXT NOT NULL,
                site TEXT,
                title TEXT,
                price_value REAL,
                price_currency TEXT,
                price_text TEXT,
                discounted_value REAL,
                discounted_formatted TEXT,
                original_formatted TEXT,
                source TEXT,
                image_url TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (history_id) REFERENCES fetch_history (id) ON DELETE CASCADE
            )
        ''')
        
        # Create indexes for better performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_history_timestamp ON fetch_history (timestamp)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_items_history_id ON items (history_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_items_url ON items (url)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_items_site ON items (site)')
    
    def _init_results_tables(self, cursor):
        """Create results dashboard tables"""
        # ===== AUTO-SCRAPER TABLES =====
        
        # Scraper runs tracking
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS scraper_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id TEXT UNIQUE NOT NULL,
                status TEXT NOT NULL DEFAULT 'running',  -- running, completed, failed, stopped
                started_at DATETIME NOT NULL,
                completed_at DATETIME,
                total_brands INTEGER DEFAULT 0,
                total_categories INTEGER DEFAULT 0,
                total_models INTEGER DEFAULT 0,
                total_products INTEGER DEFAULT 0,
                new_products INTEGER DEFAULT 0,
                updated_products INTEGER DEFAULT 0,
                errors_count INTEGER DEFAULT 0,
                current_brand TEXT,
                current_category TEXT,
                current_model TEXT,
                checkpoint TEXT,  -- JSON for resume capability
                error_log TEXT,  -- JSON array of errors
                config TEXT  -- JSON with schedule config
            )
        ''')
        
        # Brands table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ms_brands (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                slug TEXT UNIQUE NOT NULL,
                url TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Categories table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ms_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                brand_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                slug TEXT NOT NULL,
                url TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (brand_id) REFERENCES ms_brands (id) ON DELETE CASCADE,
                UNIQUE (brand_id, slug)
            )
        ''')
        
        # Models table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ms_models (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                slug TEXT NOT NULL,
                url TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (category_id) REFERENCES ms_categories (id) ON DELETE CASCADE,
                UNIQUE (category_id, slug)
            )
        ''')
        
        # Products table - the main data store
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ms_products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER NOT NULL,
                sku TEXT UNIQUE NOT NULL,
                title TEXT NOT NULL,
                description TEXT,
                price REAL,
                stock_status TEXT,  -- in_stock, out_of_stock, back_order
                availability TEXT,
                condition TEXT,  -- New, OEM, Refurbished, etc.
                product_url TEXT NOT NULL UNIQUE,
                image_urls TEXT,  -- JSON array
                variant_details TEXT,  -- JSON object (color, storage, grade, etc.)
                compatibility TEXT,  -- JSON array of compatible models
                bulk_discounts TEXT,  -- JSON object
                last_scraped_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (model_id) REFERENCES ms_models (id) ON DELETE CASCADE
            )
        ''')
        
        # Price history for tracking changes
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ms_price_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                price REAL NOT NULL,
                stock_status TEXT,
                recorded_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (product_id) REFERENCES ms_products (id) ON DELETE CASCADE
            )
        ''')
        
        # Generic product change log for non-price updates
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ms_product_changes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                change_type TEXT NOT NULL,
                old_value TEXT,
                new_value TEXT,
                metadata TEXT,
                changed_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (product_id) REFERENCES ms_products (id) ON DELETE CASCADE
            )
        ''')

        # Baseline snapshot table to capture first-seen state per product
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ms_product_baselines (
                product_id INTEGER PRIMARY KEY,
                captured_at DATETIME NOT NULL,
                snapshot TEXT,
                FOREIGN KEY (product_id) REFERENCES ms_products (id) ON DELETE CASCADE
            )
        ''')

        # Auto-scraper indexes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_scraper_runs_status ON scraper_runs (status)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_scraper_runs_started ON scraper_runs (started_at)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_brands_slug ON ms_brands (slug)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_categories_brand ON ms_categories (brand_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_models_category ON ms_models (category_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_products_model ON ms_products (model_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_products_sku ON ms_products (sku)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_products_url ON ms_products (product_url)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_price_history_product ON ms_price_history (product_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_price_history_recorded ON ms_price_history (recorded_at)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_product_changes_type ON ms_product_changes (change_type)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_product_changes_time ON ms_product_changes (changed_at)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ms_product_baselines_captured ON ms_product_baselines (captured_at)')
    
    def save_fetch_history(self, history_id: str, urls: List[str], items: List[Any], rules: Dict) -> bool:
        """Save fetch history and items to database"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Save fetch history - use the history_id as timestamp (it's already a Unix timestamp in ms)
            # Convert to Pakistan timezone for consistent storage
            timestamp_ms = int(history_id)
            timestamp_utc = datetime.datetime.fromtimestamp(timestamp_ms / 1000.0, tz=pytz.UTC)
            timestamp_pakistan = timestamp_utc.astimezone(PAKISTAN_TZ)
            
            cursor.execute('''
                INSERT INTO fetch_history (id, timestamp, urls, items_count, rules)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                history_id,
                timestamp_pakistan.isoformat(),
                json.dumps(urls),
                len(items),
                json.dumps(rules)
            ))
            
            # Save items
            for item in items:
                item_dict = asdict(item) if hasattr(item, '__dict__') else item
                cursor.execute('''
                    INSERT INTO items (
                        history_id, url, site, title, price_value, price_currency,
                        price_text, discounted_value, discounted_formatted,
                        original_formatted, source, image_url
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    history_id,
                    item_dict.get('url', ''),
                    item_dict.get('site', ''),
                    item_dict.get('title', ''),
                    item_dict.get('price_value'),
                    item_dict.get('price_currency', ''),
                    item_dict.get('price_text', ''),
                    item_dict.get('discounted_value'),
                    item_dict.get('discounted_formatted', ''),
                    item_dict.get('original_formatted', ''),
                    item_dict.get('source', ''),
                    item_dict.get('image_url', '')
                ))
            
            conn.commit()
            with self._cache_lock:
                self._brand_cache.clear()
                self._category_cache.clear()
                self._model_cache.clear()
            return True
            
        except Exception as e:
            print(f"Error saving to database: {e}")
            conn.rollback()
            return False
    
    def get_history_list(self, limit: int = 50, offset: int = 0) -> List[Dict]:
        """Get list of fetch history entries"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT id, timestamp, urls, items_count, rules, created_at
                FROM fetch_history
                ORDER BY timestamp DESC
                LIMIT ? OFFSET ?
            ''', (limit, offset))
            
            histories = []
            for row in cursor.fetchall():
                # Convert ISO timestamp back to milliseconds for frontend
                timestamp_str = row['timestamp']
                if 'T' in timestamp_str:  # ISO format
                    timestamp_dt = datetime.datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                    timestamp_ms = int(timestamp_dt.timestamp() * 1000)
                else:  # Already a number
                    timestamp_ms = int(timestamp_str)
                
                histories.append({
                    'id': row['id'],
                    'timestamp': timestamp_ms,
                    'urls': json.loads(row['urls']),
                    'items_count': row['items_count'],
                    'rules': json.loads(row['rules']),
                    'created_at': row['created_at']
                })
            
            return histories
            
        except Exception as e:
            print(f"Error getting history list: {e}")
            return []
    
    def get_history_detail(self, history_id: str) -> Optional[Dict]:
        """Get detailed history entry with items"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Get history entry
            cursor.execute('''
                SELECT id, timestamp, urls, items_count, rules, created_at
                FROM fetch_history
                WHERE id = ?
            ''', (history_id,))
            
            row = cursor.fetchone()
            if not row:
                return None
            
            # Get items for this history
            cursor.execute('''
                SELECT url, site, title, price_value, price_currency, price_text,
                       discounted_value, discounted_formatted, original_formatted,
                       source, image_url
                FROM items
                WHERE history_id = ?
                ORDER BY id
            ''', (history_id,))
            
            items = []
            for item_row in cursor.fetchall():
                items.append({
                    'url': item_row['url'],
                    'site': item_row['site'],
                    'title': item_row['title'],
                    'price_value': item_row['price_value'],
                    'price_currency': item_row['price_currency'],
                    'price_text': item_row['price_text'],
                    'discounted_value': item_row['discounted_value'],
                    'discounted_formatted': item_row['discounted_formatted'],
                    'original_formatted': item_row['original_formatted'],
                    'source': item_row['source'],
                    'image_url': item_row['image_url']
                })
            
            # Convert ISO timestamp back to milliseconds for frontend
            timestamp_str = row['timestamp']
            if 'T' in timestamp_str:  # ISO format
                timestamp_dt = datetime.datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                timestamp_ms = int(timestamp_dt.timestamp() * 1000)
            else:  # Already a number
                timestamp_ms = int(timestamp_str)
            
            return {
                'id': row['id'],
                'timestamp': timestamp_ms,
                'urls': json.loads(row['urls']),
                'items_count': row['items_count'],
                'rules': json.loads(row['rules']),
                'created_at': row['created_at'],
                'items': items
            }
            
        except Exception as e:
            print(f"Error getting history detail: {e}")
            return None
    
    def delete_history(self, history_id: str) -> bool:
        """Delete history entry and associated items"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Delete items first (due to foreign key)
            cursor.execute('DELETE FROM items WHERE history_id = ?', (history_id,))
            
            # Delete history entry
            cursor.execute('DELETE FROM fetch_history WHERE id = ?', (history_id,))
            
            conn.commit()
            return cursor.rowcount > 0
            
        except Exception as e:
            print(f"Error deleting history: {e}")
            conn.rollback()
            return False
    
    def get_statistics(self) -> Dict:
        """Get comprehensive database statistics"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Total histories
            cursor.execute('SELECT COUNT(*) as count FROM fetch_history')
            total_histories = cursor.fetchone()['count']
            
            # Total items
            cursor.execute('SELECT COUNT(*) as count FROM items')
            total_items = cursor.fetchone()['count']
            
            # Recent activity (last 30 days in Pakistan timezone)
            now_pakistan = get_pakistan_time()
            thirty_days_ago = now_pakistan - datetime.timedelta(days=30)
            thirty_days_ago_str = thirty_days_ago.isoformat()
            
            cursor.execute('''
                SELECT COUNT(*) as count FROM fetch_history
                WHERE timestamp >= ?
            ''', (thirty_days_ago_str,))
            recent_histories = cursor.fetchone()['count']
            
            # Unique models (approximation based on titles)
            cursor.execute('''
                SELECT COUNT(DISTINCT 
                    CASE 
                        WHEN title LIKE '%iPhone%' THEN 'iPhone'
                        WHEN title LIKE '%Galaxy%' THEN 'Galaxy'
                        WHEN title LIKE '%iPad%' THEN 'iPad'
                        WHEN title LIKE '%Pixel%' THEN 'Pixel'
                        WHEN title LIKE '%OnePlus%' THEN 'OnePlus'
                        ELSE SUBSTR(title, 1, 20)
                    END
                ) as unique_models FROM items WHERE title != ''
            ''')
            unique_models = cursor.fetchone()['unique_models'] or 0
            
            # Unique sites
            cursor.execute('SELECT COUNT(DISTINCT site) as count FROM items WHERE site != ""')
            unique_sites = cursor.fetchone()['count']
            
            # Database size
            cursor.execute("SELECT page_count * page_size as size FROM pragma_page_count(), pragma_page_size()")
            db_size = cursor.fetchone()['size']
            
            # Average items per session
            avg_items = round(total_items / max(total_histories, 1), 1)
            
            # Average price (from items with valid prices)
            cursor.execute('''
                SELECT AVG(price_value) as avg_price FROM items 
                WHERE price_value IS NOT NULL AND price_value > 0
            ''')
            avg_price_result = cursor.fetchone()
            avg_price = round(avg_price_result['avg_price'] or 0, 2)
            
            # Success rate (items with prices vs total items)
            cursor.execute('''
                SELECT 
                    COUNT(CASE WHEN price_value IS NOT NULL AND price_value > 0 THEN 1 END) as successful,
                    COUNT(*) as total
                FROM items
            ''')
            success_data = cursor.fetchone()
            success_rate = round((success_data['successful'] / max(success_data['total'], 1)) * 100, 1)
            
            # Top site by item count
            cursor.execute('''
                SELECT site, COUNT(*) as item_count 
                FROM items 
                WHERE site != "" 
                GROUP BY site 
                ORDER BY item_count DESC 
                LIMIT 1
            ''')
            top_site_result = cursor.fetchone()
            top_site = top_site_result['site'] if top_site_result else 'N/A'
            
            # Clean up site name for display
            if top_site and top_site != 'N/A':
                # Remove common prefixes and make it shorter
                top_site = top_site.replace('www.', '').replace('.com', '').replace('.ca', '')
                if '.' in top_site:
                    top_site = top_site.split('.')[0]
                top_site = top_site.capitalize()
            
            # Latest session date
            cursor.execute('''
                SELECT timestamp 
                FROM fetch_history 
                ORDER BY timestamp DESC 
                LIMIT 1
            ''')
            latest_session_result = cursor.fetchone()
            latest_session = 'Never'
            if latest_session_result:
                try:
                    # Parse the timestamp and convert to Pakistan time
                    ts_str = latest_session_result['timestamp']
                    if '+' in ts_str or 'Z' in ts_str:
                        # Has timezone info
                        ts = datetime.datetime.fromisoformat(ts_str.replace('Z', '+00:00'))
                        if ts.tzinfo is None:
                            ts = pytz.UTC.localize(ts)
                        ts_pakistan = ts.astimezone(PAKISTAN_TZ)
                    else:
                        # Assume it's already in Pakistan time
                        ts = datetime.datetime.fromisoformat(ts_str)
                        ts_pakistan = PAKISTAN_TZ.localize(ts)
                    
                    latest_session = ts_pakistan.strftime('%b %d')  # e.g., "Oct 17"
                except Exception as e:
                    print(f"Error parsing latest session timestamp: {e}")
                    latest_session = 'Recent'
            
            # Oldest session date
            cursor.execute('''
                SELECT timestamp 
                FROM fetch_history 
                ORDER BY timestamp ASC 
                LIMIT 1
            ''')
            oldest_session_result = cursor.fetchone()
            oldest_session = 'N/A'
            if oldest_session_result:
                try:
                    # Parse the timestamp and convert to Pakistan time
                    ts_str = oldest_session_result['timestamp']
                    if '+' in ts_str or 'Z' in ts_str:
                        # Has timezone info
                        ts = datetime.datetime.fromisoformat(ts_str.replace('Z', '+00:00'))
                        if ts.tzinfo is None:
                            ts = pytz.UTC.localize(ts)
                        ts_pakistan = ts.astimezone(PAKISTAN_TZ)
                    else:
                        # Assume it's already in Pakistan time
                        ts = datetime.datetime.fromisoformat(ts_str)
                        ts_pakistan = PAKISTAN_TZ.localize(ts)
                    
                    oldest_session = ts_pakistan.strftime('%b %d, %Y')  # e.g., "Oct 17, 2024"
                except Exception as e:
                    print(f"Error parsing oldest session timestamp: {e}")
                    oldest_session = 'Unknown'
            
            # Total value of all items
            cursor.execute('''
                SELECT SUM(price_value) as total_value FROM items 
                WHERE price_value IS NOT NULL AND price_value > 0
            ''')
            total_value_result = cursor.fetchone()
            total_value = round(total_value_result['total_value'] or 0, 2)
            
            # Highest price
            cursor.execute('''
                SELECT MAX(price_value) as highest_price FROM items 
                WHERE price_value IS NOT NULL AND price_value > 0
            ''')
            highest_price_result = cursor.fetchone()
            highest_price = round(highest_price_result['highest_price'] or 0, 2)
            
            # Lowest price
            cursor.execute('''
                SELECT MIN(price_value) as lowest_price FROM items 
                WHERE price_value IS NOT NULL AND price_value > 0
            ''')
            lowest_price_result = cursor.fetchone()
            lowest_price = round(lowest_price_result['lowest_price'] or 0, 2)
            
            return {
                'total_histories': total_histories,
                'total_items': total_items,
                'recent_histories': recent_histories,
                'unique_models': unique_models,
                'unique_sites': unique_sites,
                'database_size': db_size,
                'avg_items_per_session': avg_items,
                'avg_price': avg_price,
                'success_rate': success_rate,
                'top_site': top_site,
                'latest_session': latest_session,
                'oldest_session': oldest_session,
                'total_value': total_value,
                'highest_price': highest_price,
                'lowest_price': lowest_price
            }
            
        except Exception as e:
            print(f"Error getting statistics: {e}")
            return {}
    
    
    def cleanup_old_entries(self, days: int = 90) -> int:
        """Remove entries older than specified days (calculated in Pakistan timezone)"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # If days is very large (99999), delete everything
            if days >= 99999:
                # Count total entries before deletion
                cursor.execute('SELECT COUNT(*) as count FROM fetch_history')
                total_count = cursor.fetchone()['count']
                
                # Delete all items
                cursor.execute('DELETE FROM items')
                
                # Delete all history
                cursor.execute('DELETE FROM fetch_history')
                
                conn.commit()
                return total_count
            
            # Calculate cutoff date in Pakistan timezone
            now_pakistan = get_pakistan_time()
            cutoff_date = now_pakistan - datetime.timedelta(days=days)
            cutoff_date_str = cutoff_date.isoformat()
            
            # Get old history IDs
            cursor.execute('''
                SELECT id FROM fetch_history
                WHERE timestamp < ?
            ''', (cutoff_date_str,))
            
            old_ids = [row['id'] for row in cursor.fetchall()]
            
            if old_ids:
                # Delete items for old histories
                placeholders = ','.join(['?' for _ in old_ids])
                cursor.execute(f'DELETE FROM items WHERE history_id IN ({placeholders})', old_ids)
                
                # Delete old histories
                cursor.execute(f'DELETE FROM fetch_history WHERE id IN ({placeholders})', old_ids)
                
                conn.commit()
                return len(old_ids)
            
            return 0
            
        except Exception as e:
            print(f"Error cleaning up old entries: {e}")
            conn.rollback()
            return 0
    
    def search_items(self, query: str, limit: int = 100) -> List[Dict]:
        """Search items by title or URL"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT i.*, h.timestamp
                FROM items i
                JOIN fetch_history h ON i.history_id = h.id
                WHERE i.title LIKE ? OR i.url LIKE ?
                ORDER BY h.timestamp DESC
                LIMIT ?
            ''', (f'%{query}%', f'%{query}%', limit))
            
            items = []
            for row in cursor.fetchall():
                items.append({
                    'url': row['url'],
                    'site': row['site'],
                    'title': row['title'],
                    'price_value': row['price_value'],
                    'price_currency': row['price_currency'],
                    'price_text': row['price_text'],
                    'discounted_value': row['discounted_value'],
                    'discounted_formatted': row['discounted_formatted'],
                    'original_formatted': row['original_formatted'],
                    'source': row['source'],
                    'image_url': row['image_url'],
                    'timestamp': row['timestamp'],
                    'history_id': row['history_id']
                })
            
            return items
            
        except Exception as e:
            print(f"Error searching items: {e}")
            return []

    # ===== AUTO-SCRAPER METHODS =====
    
    def create_scraper_run(self, run_id: str, config: Dict = None) -> bool:
        """Create a new scraper run entry"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO scraper_runs (run_id, status, started_at, config)
                VALUES (?, 'running', ?, ?)
            ''', (run_id, datetime.datetime.now(), json.dumps(config or {})))
            
            conn.commit()
            return True
        except Exception as e:
            print(f"Error creating scraper run: {e}")
            return False
    
    def update_scraper_run(self, run_id: str, updates: Dict) -> bool:
        """Update scraper run with progress or completion"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            set_clauses = []
            values = []
            
            for key, value in updates.items():
                if key in ['status', 'completed_at', 'total_brands', 'total_categories', 
                           'total_models', 'total_products', 'new_products', 'updated_products',
                           'errors_count', 'current_brand', 'current_category', 'current_model']:
                    set_clauses.append(f"{key} = ?")
                    values.append(value)
                elif key in ['checkpoint', 'error_log']:
                    set_clauses.append(f"{key} = ?")
                    values.append(json.dumps(value) if value else None)
            
            if not set_clauses:
                return False
            
            values.append(run_id)
            query = f"UPDATE scraper_runs SET {', '.join(set_clauses)} WHERE run_id = ?"
            
            cursor.execute(query, values)
            conn.commit()
            return cursor.rowcount > 0
            
        except Exception as e:
            print(f"Error updating scraper run: {e}")
            return False
    
    def get_scraper_run(self, run_id: str) -> Optional[Dict]:
        """Get scraper run details"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT * FROM scraper_runs WHERE run_id = ?
            ''', (run_id,))
            
            row = cursor.fetchone()
            if not row:
                return None
            
            return {
                'id': row['id'],
                'run_id': row['run_id'],
                'status': row['status'],
                'started_at': row['started_at'],
                'completed_at': row['completed_at'],
                'total_brands': row['total_brands'],
                'total_categories': row['total_categories'],
                'total_models': row['total_models'],
                'total_products': row['total_products'],
                'new_products': row['new_products'],
                'updated_products': row['updated_products'],
                'errors_count': row['errors_count'],
                'current_brand': row['current_brand'],
                'current_category': row['current_category'],
                'current_model': row['current_model'],
                'checkpoint': json.loads(row['checkpoint']) if row['checkpoint'] else None,
                'error_log': json.loads(row['error_log']) if row['error_log'] else [],
                'config': json.loads(row['config']) if row['config'] else {}
            }
            
        except Exception as e:
            print(f"Error getting scraper run: {e}")
            return None

    # ===== Maintenance helpers =====

    def clear_all_data(self, vacuum: bool = True) -> Dict[str, int]:
        """Delete all records from managed tables. Returns counts per table."""
        conn = self.get_connection()
        cursor = conn.cursor()
        summary: Dict[str, int] = {}

        try:
            if self.create_results_tables:
                results_tables = [
                    'ms_price_history',
                    'ms_product_changes',
                    'ms_products',
                    'ms_models',
                    'ms_categories',
                    'ms_brands',
                    'scraper_runs'
                ]
                for table in results_tables:
                    cursor.execute(f'SELECT COUNT(*) FROM {table}')
                    count = cursor.fetchone()[0]
                    cursor.execute(f'DELETE FROM {table}')
                    summary[table] = count

            if self.create_legacy_tables:
                legacy_tables = ['items', 'fetch_history']
                for table in legacy_tables:
                    cursor.execute(f'SELECT COUNT(*) FROM {table}')
                    count = cursor.fetchone()[0]
                    cursor.execute(f'DELETE FROM {table}')
                    summary[table] = count

            # Reset autoincrement sequences for cleaned tables
            tables_to_reset = []
            if self.create_results_tables:
                tables_to_reset.extend(['ms_price_history', 'ms_product_changes', 'ms_products', 'ms_models', 'ms_categories', 'ms_brands', 'scraper_runs'])
            if self.create_legacy_tables:
                tables_to_reset.extend(['items', 'fetch_history'])

            if tables_to_reset:
                placeholders = ','.join('?' for _ in tables_to_reset)
                try:
                    cursor.execute(f"DELETE FROM sqlite_sequence WHERE name IN ({placeholders})", tables_to_reset)
                except sqlite3.OperationalError:
                    # sqlite_sequence may not exist if AUTOINCREMENT was never used
                    pass

            conn.commit()

        except Exception as exc:
            conn.rollback()
            raise exc
        finally:
            if vacuum:
                try:
                    cursor.execute('VACUUM')
                except Exception:
                    pass

        return summary
    
    def get_scraper_runs_list(self, limit: int = 20) -> List[Dict]:
        """Get list of recent scraper runs"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT id, run_id, status, started_at, completed_at,
                       total_brands, total_categories, total_models, total_products,
                       new_products, updated_products, errors_count
                FROM scraper_runs
                ORDER BY started_at DESC
                LIMIT ?
            ''', (limit,))
            
            runs = []
            for row in cursor.fetchall():
                runs.append({
                    'id': row['id'],
                    'run_id': row['run_id'],
                    'status': row['status'],
                    'started_at': row['started_at'],
                    'completed_at': row['completed_at'],
                    'total_brands': row['total_brands'],
                    'total_categories': row['total_categories'],
                    'total_models': row['total_models'],
                    'total_products': row['total_products'],
                    'new_products': row['new_products'],
                    'updated_products': row['updated_products'],
                    'errors_count': row['errors_count']
                })
            
            return runs
            
        except Exception as e:
            print(f"Error getting scraper runs list: {e}")
            return []
    
    def save_brand(self, name: str, slug: str, url: str) -> int:
        """Save or update brand, returns brand_id"""
        try:
            cache_key = (slug or '').strip().lower()
            with self._cache_lock:
                cached_id = self._brand_cache.get(cache_key)
            if cached_id:
                return cached_id

            conn = self.get_connection()
            cursor = conn.cursor()

            cursor.execute('SELECT id, name, url FROM ms_brands WHERE slug = ?', (slug,))
            row = cursor.fetchone()
            now = datetime.datetime.now()

            if row:
                brand_id = row['id']
                if row['name'] != name or row['url'] != url:
                    cursor.execute('''
                        UPDATE ms_brands SET name = ?, url = ?, updated_at = ?
                        WHERE id = ?
                    ''', (name, url, now, brand_id))
                    conn.commit()
                with self._cache_lock:
                    self._brand_cache[cache_key] = brand_id
                return brand_id

            cursor.execute('''
                INSERT INTO ms_brands (name, slug, url, updated_at)
                VALUES (?, ?, ?, ?)
            ''', (name, slug, url, now))
            conn.commit()
            brand_id = cursor.lastrowid

            with self._cache_lock:
                self._brand_cache[cache_key] = brand_id

            return brand_id
            
        except Exception as e:
            print(f"Error saving brand: {e}")
            return 0
    
    def save_category(self, brand_id: int, name: str, slug: str, url: str) -> int:
        """Save or update category, returns category_id"""
        try:
            cache_key = (brand_id, (slug or '').strip().lower())
            with self._cache_lock:
                cached_id = self._category_cache.get(cache_key)
            if cached_id:
                return cached_id

            conn = self.get_connection()
            cursor = conn.cursor()

            cursor.execute('SELECT id, name, url FROM ms_categories WHERE brand_id = ? AND slug = ?', (brand_id, slug))
            row = cursor.fetchone()
            now = datetime.datetime.now()

            if row:
                category_id = row['id']
                if row['name'] != name or row['url'] != url:
                    cursor.execute('''
                        UPDATE ms_categories SET name = ?, url = ?, updated_at = ?
                        WHERE id = ?
                    ''', (name, url, now, category_id))
                    conn.commit()
                with self._cache_lock:
                    self._category_cache[cache_key] = category_id
                return category_id

            cursor.execute('''
                INSERT INTO ms_categories (brand_id, name, slug, url, updated_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (brand_id, name, slug, url, now))
            conn.commit()
            category_id = cursor.lastrowid

            with self._cache_lock:
                self._category_cache[cache_key] = category_id

            return category_id
            
        except Exception as e:
            print(f"Error saving category: {e}")
            return 0
    
    def save_model(self, category_id: int, name: str, slug: str, url: str) -> int:
        """Save or update model, returns model_id"""
        try:
            cache_key = (category_id, (slug or '').strip().lower())
            with self._cache_lock:
                cached_id = self._model_cache.get(cache_key)
            if cached_id:
                return cached_id

            conn = self.get_connection()
            cursor = conn.cursor()

            cursor.execute('SELECT id, name, url FROM ms_models WHERE category_id = ? AND slug = ?', (category_id, slug))
            row = cursor.fetchone()
            now = datetime.datetime.now()

            if row:
                model_id = row['id']
                if row['name'] != name or row['url'] != url:
                    cursor.execute('''
                        UPDATE ms_models SET name = ?, url = ?, updated_at = ?
                        WHERE id = ?
                    ''', (name, url, now, model_id))
                    conn.commit()
                with self._cache_lock:
                    self._model_cache[cache_key] = model_id
                return model_id

            cursor.execute('''
                INSERT INTO ms_models (category_id, name, slug, url, updated_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (category_id, name, slug, url, now))
            conn.commit()
            model_id = cursor.lastrowid

            with self._cache_lock:
                self._model_cache[cache_key] = model_id

            return model_id
            
        except Exception as e:
            print(f"Error saving model: {e}")
            return 0
    
    def _truncate_change_value(self, value: Optional[str], limit: int = 500) -> Optional[str]:
        if value is None:
            return None
        trimmed = value.strip()
        if not trimmed:
            return None
        if len(trimmed) <= limit:
            return trimmed
        return trimmed[:limit - 3] + '...'

    def _normalize_change_value_for_compare(self, value: Optional[str]) -> str:
        if value is None:
            return ''

        cleaned = str(value).strip()
        if not cleaned:
            return ''

        cleaned = re.sub(r'[\s_-]+', ' ', cleaned)
        return cleaned.lower()

    def _has_meaningful_change(
        self,
        change_type: Optional[str],
        old_value: Optional[str],
        new_value: Optional[str]
    ) -> bool:
        normalized_type = (change_type or '').strip().lower()

        old_norm = self._normalize_change_value_for_compare(old_value)
        new_norm = self._normalize_change_value_for_compare(new_value)

        if not old_norm and not new_norm:
            return False

        if normalized_type == 'description':
            old_trimmed = (old_value or '').strip()
            new_trimmed = (new_value or '').strip()
            if old_trimmed == new_trimmed:
                return False

        return old_norm != new_norm

    def _build_change_type_case_sql(self, column: str) -> str:
        sanitized_column = column.strip()
        return (
            "CASE "
            f"WHEN LOWER({sanitized_column}) IN ('stock', 'stock_status', 'availability', 'inventory') "
            f"    OR LOWER({sanitized_column}) LIKE '%stock%' "
            f"    OR LOWER({sanitized_column}) LIKE '%availability%' "
            f"    OR LOWER({sanitized_column}) LIKE '%inventory%' "
            f"    OR LOWER({sanitized_column}) LIKE '%qty%' THEN 'stock' "
            f"WHEN LOWER({sanitized_column}) IN ('description', 'desc', 'details', 'product_description') "
            f"    OR LOWER({sanitized_column}) LIKE '%description%' "
            f"    OR LOWER({sanitized_column}) LIKE '%content%' "
            f"    OR LOWER({sanitized_column}) LIKE '%copy%' THEN 'description' "
            f"WHEN LOWER({sanitized_column}) IN ('price', 'pricing', 'cost') "
            f"    OR LOWER({sanitized_column}) LIKE '%price%' "
            f"    OR LOWER({sanitized_column}) LIKE '%cost%' THEN 'price' "
            f"ELSE LOWER({sanitized_column}) END"
        )

    def _build_significant_change_predicate(self, table_alias: str) -> str:
        alias = table_alias.strip()
        return (
            f"(({alias}.old_value IS NULL AND {alias}.new_value IS NOT NULL) "
            f" OR ({alias}.old_value IS NOT NULL AND {alias}.new_value IS NULL) "
            f" OR ({alias}.old_value IS NOT NULL AND {alias}.new_value IS NOT NULL "
            f"     AND TRIM(LOWER({alias}.old_value)) != TRIM(LOWER({alias}.new_value))))"
        )

    def _build_baseline_exclusion_predicate(self, change_alias: str, product_alias: str) -> str:
        c_alias = change_alias.strip()
        p_alias = product_alias.strip()
        return (
            f"(({c_alias}.old_value IS NULL OR TRIM({c_alias}.old_value) = '') "
            f" AND ({c_alias}.new_value IS NOT NULL AND TRIM({c_alias}.new_value) != '') "
            f" AND datetime({c_alias}.changed_at) <= datetime({p_alias}.created_at))"
        )

    def _log_product_change(
        self,
        product_id: int,
        change_type: str,
        old_value: Optional[str],
        new_value: Optional[str],
        changed_at: Optional[datetime.datetime] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> None:
        try:
            if not self._has_meaningful_change(change_type, old_value, new_value):
                return

            conn = self.get_connection()
            cursor = conn.cursor()
            payload = json.dumps(metadata, ensure_ascii=False) if metadata else None
            cursor.execute('''
                INSERT INTO ms_product_changes (product_id, change_type, old_value, new_value, metadata, changed_at)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                product_id,
                change_type,
                self._truncate_change_value(old_value),
                self._truncate_change_value(new_value),
                payload,
                changed_at or datetime.datetime.now()
            ))
        except Exception as log_error:
            print(f"Error logging product change ({change_type}): {log_error}")

    def _format_change_value(self, change_type: str, numeric_value: Optional[float], text_value: Optional[str]) -> str:
        if change_type == 'price':
            if numeric_value is None:
                return '—'
            return f"${numeric_value:.2f}"

        text = (text_value or '').strip()
        if not text:
            return '—'

        if change_type == 'stock':
            cleaned = re.sub(r'[_-]+', ' ', text)
            return cleaned.title()

        return text

    def _format_change_delta_text(self, change_type: str, delta: Optional[float]) -> Optional[str]:
        if change_type != 'price' or delta is None:
            return None

        if not isinstance(delta, (int, float)):
            return None

        if abs(delta) < 1e-9:
            return None

        prefix = '+' if delta > 0 else '-'
        return f"{prefix}${abs(delta):.2f}"

    def _build_change_label(
        self,
        change_type: str,
        old_display: str,
        new_display: str,
        delta: Optional[float]
    ) -> str:
        normalized = (change_type or '').strip().lower()

        if normalized == 'price':
            if isinstance(delta, (int, float)) and abs(delta) >= 1e-9:
                return 'Price increased' if delta > 0 else 'Price decreased'
            return 'Price changed'

        if normalized == 'stock':
            return 'Stock status changed'

        if normalized == 'description':
            return 'Description updated'

        return 'Value updated'

    def _sanitize_product_title(self, title: Optional[str]) -> str:
        cleaned = (title or '').strip()
        return cleaned if cleaned else 'Untitled product'

    def _clean_category_name(self, category: Optional[str]) -> str:
        if not category:
            return ''
        cleaned = re.sub(r'\s+', ' ', category).strip()
        return cleaned

    def _safe_load_json(self, payload: Optional[str], default):
        if not payload:
            return default
        try:
            return json.loads(payload)
        except Exception:
            return default

    def _build_model_identifier(self, brand: Optional[str], model: Optional[str], compatibility: Optional[List[Any]]) -> str:
        parts = [part for part in [brand, model] if part]
        identifier = ' '.join(parts).strip()
        if identifier:
            return identifier

        if compatibility:
            for entry in compatibility:
                if isinstance(entry, str) and entry.strip():
                    return entry.strip()

        return ''

    def _extract_chipset(self, variant_details: Optional[Dict[str, Any]], metadata: Optional[Dict[str, Any]], title: Optional[str]) -> str:
        candidates: List[str] = []

        for source in (variant_details, metadata):
            if isinstance(source, dict):
                for key in ('chipset', 'chip', 'processor'):
                    value = source.get(key)
                    if isinstance(value, str) and value.strip():
                        return value.strip()
                    if value:
                        return str(value)

        if title:
            match = re.search(r'(A\d{2}|Snapdragon\s?\d+|Dimensity\s?\d+)', title, re.IGNORECASE)
            if match:
                return match.group(1)

        return ''

    def save_product(self, product_data: Dict) -> Tuple[int, bool]:
        """
        Save or update product, returns (product_id, is_new)
        is_new indicates if this was a new insert or an update
        """
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Check if product exists
            cursor.execute('SELECT id, price, stock_status, description FROM ms_products WHERE sku = ?', 
                          (product_data['sku'],))
            existing = cursor.fetchone()
            
            is_new = existing is None
            now = datetime.datetime.now()
            
            if is_new:
                # Insert new product
                cursor.execute('''
                    INSERT INTO ms_products (
                        model_id, sku, title, description, price, stock_status,
                        availability, condition, product_url, image_urls,
                        variant_details, compatibility, bulk_discounts,
                        last_scraped_at, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    product_data['model_id'],
                    product_data['sku'],
                    product_data['title'],
                    product_data.get('description', ''),
                    product_data.get('price'),
                    product_data.get('stock_status', ''),
                    product_data.get('availability', ''),
                    product_data.get('condition', ''),
                    product_data['product_url'],
                    json.dumps(product_data.get('image_urls', [])),
                    json.dumps(product_data.get('variant_details', {})),
                    json.dumps(product_data.get('compatibility', [])),
                    json.dumps(product_data.get('bulk_discounts', {})),
                    now, now, now
                ))
                
                product_id = cursor.lastrowid

                if product_data.get('price') is not None:
                    cursor.execute('''
                        INSERT INTO ms_price_history (product_id, price, stock_status, recorded_at)
                        VALUES (?, ?, ?, ?)
                    ''', (
                        product_id,
                        product_data.get('price'),
                        product_data.get('stock_status', ''),
                        now
                    ))

                baseline_payload = {
                    "sku": product_data['sku'],
                    "title": product_data.get('title', ''),
                    "description": product_data.get('description', ''),
                    "price": product_data.get('price'),
                    "stock_status": product_data.get('stock_status', ''),
                    "model_id": product_data['model_id']
                }

                cursor.execute('''
                    INSERT OR IGNORE INTO ms_product_baselines (product_id, captured_at, snapshot)
                    VALUES (?, ?, ?)
                ''', (
                    product_id,
                    now,
                    json.dumps(baseline_payload, ensure_ascii=False)
                ))

            else:
                # Update existing product
                product_id = existing['id']
                previous_price = existing['price']
                previous_stock = existing['stock_status'] or ''
                previous_description = existing['description'] or ''

                cursor.execute('''
                    UPDATE ms_products SET
                        title = ?, description = ?, price = ?, stock_status = ?,
                        availability = ?, condition = ?, image_urls = ?,
                        variant_details = ?, compatibility = ?, bulk_discounts = ?,
                        last_scraped_at = ?, updated_at = ?
                    WHERE id = ?
                ''', (
                    product_data['title'],
                    product_data.get('description', ''),
                    product_data.get('price'),
                    product_data.get('stock_status', ''),
                    product_data.get('availability', ''),
                    product_data.get('condition', ''),
                    json.dumps(product_data.get('image_urls', [])),
                    json.dumps(product_data.get('variant_details', {})),
                    json.dumps(product_data.get('compatibility', [])),
                    json.dumps(product_data.get('bulk_discounts', {})),
                    now, now, product_id
                ))
                
                # Track price change
                if existing and product_data.get('price') != previous_price:
                    cursor.execute('''
                        INSERT INTO ms_price_history (product_id, price, stock_status, recorded_at)
                        VALUES (?, ?, ?, ?)
                    ''', (product_id, product_data.get('price'), 
                          product_data.get('stock_status', ''), now))

                if existing:
                    current_stock = (product_data.get('stock_status') or '').strip()
                    previous_stock_clean = (previous_stock or '').strip()
                    if previous_stock_clean.lower() != current_stock.lower():
                        if current_stock:
                            self._log_product_change(
                                product_id,
                                'stock',
                                old_value=previous_stock,
                                new_value=product_data.get('stock_status'),
                                changed_at=now
                            )

                    new_description = (product_data.get('description') or '').strip()
                    if new_description and new_description != previous_description.strip():
                        self._log_product_change(
                            product_id,
                            'description',
                            old_value=previous_description,
                            new_value=product_data.get('description'),
                            changed_at=now
                        )
            
            conn.commit()
            return (product_id, is_new)
            
        except Exception as e:
            print(f"Error saving product: {e}")
            return (0, False)
    
    def get_scraper_statistics(self) -> Dict:
        """Get comprehensive auto-scraper statistics"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            stats = {}
            
            # Total counts
            cursor.execute('SELECT COUNT(*) as count FROM ms_brands')
            stats['total_brands'] = cursor.fetchone()['count']
            
            cursor.execute('SELECT COUNT(*) as count FROM ms_categories')
            stats['total_categories'] = cursor.fetchone()['count']
            
            cursor.execute('SELECT COUNT(*) as count FROM ms_models')
            stats['total_models'] = cursor.fetchone()['count']
            
            cursor.execute('SELECT COUNT(*) as count FROM ms_products')
            stats['total_products'] = cursor.fetchone()['count']
            
            # Recent runs
            cursor.execute('''
                SELECT COUNT(*) as count FROM scraper_runs
                WHERE started_at >= datetime('now', '-7 days')
            ''')
            stats['runs_last_7_days'] = cursor.fetchone()['count']
            
            # Last run info
            cursor.execute('''
                SELECT status, started_at, completed_at, total_products
                FROM scraper_runs
                ORDER BY started_at DESC
                LIMIT 1
            ''')
            last_run = cursor.fetchone()
            if last_run:
                stats['last_run_status'] = last_run['status']
                stats['last_run_date'] = last_run['started_at']
                stats['last_run_products'] = last_run['total_products']
            
            # Products with price changes (last 7 days)
            cursor.execute('''
                SELECT COUNT(DISTINCT product_id) as count
                FROM ms_price_history
                WHERE recorded_at >= datetime('now', '-7 days')
            ''')
            stats['price_changes_7_days'] = cursor.fetchone()['count']
            
            # Average price
            cursor.execute('SELECT AVG(price) as avg FROM ms_products WHERE price > 0')
            avg_result = cursor.fetchone()
            stats['avg_price'] = round(avg_result['avg'] or 0, 2)
            
            # In stock products
            cursor.execute('''
                SELECT COUNT(*) as count FROM ms_products
                WHERE stock_status = 'in_stock'
            ''')
            stats['in_stock_products'] = cursor.fetchone()['count']
            
            return stats
            
        except Exception as e:
            print(f"Error getting scraper statistics: {e}")
            return {}
    
    def search_products(self, query: str = '', brand: str = '', category: str = '', 
                       model: str = '', limit: int = 100) -> List[Dict]:
        """Search products with filters"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            sql = '''
                SELECT 
                    p.*,
                    m.name as model_name,
                    c.name as category_name,
                    b.name as brand_name
                FROM ms_products p
                JOIN ms_models m ON p.model_id = m.id
                JOIN ms_categories c ON m.category_id = c.id
                JOIN ms_brands b ON c.brand_id = b.id
                WHERE 1=1
            '''
            params = []
            
            if query:
                sql += ' AND (p.title LIKE ? OR p.description LIKE ? OR p.sku LIKE ?)'
                params.extend([f'%{query}%', f'%{query}%', f'%{query}%'])
            
            if brand:
                sql += ' AND b.slug = ?'
                params.append(brand)
            
            if category:
                sql += ' AND c.slug = ?'
                params.append(category)
            
            if model:
                sql += ' AND m.slug = ?'
                params.append(model)
            
            sql += ' ORDER BY p.updated_at DESC LIMIT ?'
            params.append(limit)
            
            cursor.execute(sql, params)
            
            products = []
            for row in cursor.fetchall():
                products.append({
                    'id': row['id'],
                    'sku': row['sku'],
                    'title': row['title'],
                    'description': row['description'],
                    'price': row['price'],
                    'stock_status': row['stock_status'],
                    'availability': row['availability'],
                    'condition': row['condition'],
                    'product_url': row['product_url'],
                    'image_urls': json.loads(row['image_urls']) if row['image_urls'] else [],
                    'brand_name': row['brand_name'],
                    'category_name': row['category_name'],
                    'model_name': row['model_name'],
                    'updated_at': row['updated_at']
                })
            
            return products
            
        except Exception as e:
            print(f"Error searching products: {e}")
            return []

    # ===== RESULTS DASHBOARD HELPERS =====
    
    def get_clients(self) -> List[str]:
        """Get list of available client sites"""
        return ["mobilesentrix", "xcellparts", "txparts"]
    
    def get_totals(self, client: str) -> Dict[str, int]:
        """Get total counts and 24h changes for a client"""
        try:
            conn = self.get_connection()
            conn.execute('PRAGMA journal_mode=WAL')
            conn.execute('PRAGMA busy_timeout=5000')
            cursor = conn.cursor()
            
            totals = {
                "total_products": 0,
                "price_changes_24h": 0,
                "stock_changes_24h": 0,
                "description_updates_24h": 0
            }
            
            if client == "mobilesentrix":
                # Total products from ms_products
                cursor.execute('SELECT COUNT(*) as count FROM ms_products')
                totals["total_products"] = cursor.fetchone()["count"] or 0
                
                # Price change events (exclude baseline inserts)
                cursor.execute('''
                    SELECT COUNT(*) as count
                    FROM ms_price_history ph
                    JOIN ms_price_history prev
                        ON prev.product_id = ph.product_id
                       AND prev.recorded_at = (
                            SELECT MAX(recorded_at)
                            FROM ms_price_history
                            WHERE product_id = ph.product_id
                              AND recorded_at < ph.recorded_at
                        )
                    WHERE ph.recorded_at >= datetime('now', '-1 day')
                      AND (
                          (prev.price IS NULL AND ph.price IS NOT NULL)
                          OR (prev.price IS NOT NULL AND ph.price IS NULL)
                          OR (
                              prev.price IS NOT NULL
                              AND ph.price IS NOT NULL
                              AND prev.price != ph.price
                          )
                      )
                ''')
                price_row = cursor.fetchone()
                totals["price_changes_24h"] = price_row["count"] if price_row else 0

                normalization_case = self._build_change_type_case_sql('pc.change_type')
                significant_predicate = self._build_significant_change_predicate('pc')
                baseline_exclusion = self._build_baseline_exclusion_predicate('pc', 'p')

                cursor.execute(f'''
                    SELECT
                        COALESCE(SUM(CASE WHEN normalized_type = 'stock' THEN 1 ELSE 0 END), 0) AS stock_changes,
                        COALESCE(SUM(CASE WHEN normalized_type = 'description' THEN 1 ELSE 0 END), 0) AS description_changes
                    FROM (
                        SELECT {normalization_case} AS normalized_type
                        FROM ms_product_changes pc
                        JOIN ms_products p ON p.id = pc.product_id
                        WHERE pc.changed_at >= datetime('now', '-1 day')
                          AND {significant_predicate}
                          AND NOT {baseline_exclusion}
                    ) filtered
                ''')
                change_row = cursor.fetchone()
                totals["stock_changes_24h"] = (change_row["stock_changes"] if change_row and change_row["stock_changes"] is not None else 0)
                totals["description_updates_24h"] = (change_row["description_changes"] if change_row and change_row["description_changes"] is not None else 0)
                
            elif client == "xcellparts":
                # For XCell, use items table with site filter
                cursor.execute('SELECT COUNT(*) as count FROM items WHERE site LIKE ?', ('%xcell%',))
                totals["total_products"] = cursor.fetchone()["count"] or 0
                
                # For changes, we'll approximate with recent items
                cursor.execute('''
                    SELECT COUNT(*) as count FROM items 
                    WHERE site LIKE ? AND created_at >= datetime('now', '-1 day')
                ''', ('%xcell%',))
                recent_count = cursor.fetchone()["count"] or 0
                totals["price_changes_24h"] = recent_count
                totals["stock_changes_24h"] = recent_count

            elif client == "txparts":
                cursor.execute('SELECT COUNT(*) as count FROM items WHERE site LIKE ?', ('%txparts%',))
                totals["total_products"] = cursor.fetchone()["count"] or 0

                cursor.execute('''
                    SELECT COUNT(*) as count FROM items 
                    WHERE site LIKE ? AND created_at >= datetime('now', '-1 day')
                ''', ('%txparts%',))
                recent_count = cursor.fetchone()["count"] or 0
                totals["price_changes_24h"] = recent_count
                totals["stock_changes_24h"] = recent_count
            
            return totals
            
        except Exception as e:
            print(f"Error getting totals for {client}: {e}")
            return {"total_products": 0, "price_changes_24h": 0, "stock_changes_24h": 0, "description_updates_24h": 0}
    
    def get_category_completion(self, client: str) -> Dict[str, any]:
        """Get category discovery and completion stats"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            if client == "mobilesentrix":
                # Count total categories discovered
                cursor.execute('SELECT COUNT(*) as total FROM ms_categories')
                total = cursor.fetchone()["total"] or 0
                
                # Count categories with products (completed)
                cursor.execute('''
                    SELECT COUNT(DISTINCT c.id) as completed
                    FROM ms_categories c
                    JOIN ms_models m ON c.id = m.category_id
                    JOIN ms_products p ON m.id = p.model_id
                ''')
                completed = cursor.fetchone()["completed"] or 0
                
            elif client == "xcellparts":
                site_key = 'xcellparts'
                # Estimate based on distinct category paths stored in legacy rules
                cursor.execute('''
                    SELECT COUNT(DISTINCT json_extract(rules, '$.category_path')) as total
                    FROM fetch_history
                    WHERE json_extract(rules, '$.site') = ?
                ''', (site_key,))
                total = cursor.fetchone()["total"] or 1  # Avoid division by zero
                
                cursor.execute('''
                    SELECT COUNT(DISTINCT json_extract(rules, '$.category_path')) as completed
                    FROM fetch_history h
                    JOIN items i ON h.id = i.history_id
                    WHERE json_extract(h.rules, '$.site') = ?
                ''', (site_key,))
                completed = cursor.fetchone()["completed"] or 0

            elif client == "txparts":
                cursor.execute('''
                    SELECT COUNT(DISTINCT json_each.value) as total
                    FROM fetch_history, json_each(fetch_history.rules, '$.category_urls')
                    WHERE json_extract(fetch_history.rules, '$.site') = 'txparts'
                ''')
                total = cursor.fetchone()["total"] or 1

                cursor.execute('''
                    SELECT COUNT(DISTINCT json_each.value) as completed
                    FROM fetch_history h
                    JOIN items i ON h.id = i.history_id
                    JOIN json_each(h.rules, '$.category_urls')
                    WHERE json_extract(h.rules, '$.site') = 'txparts'
                ''')
                completed = cursor.fetchone()["completed"] or 0

            else:
                total = 0
                completed = 0
            
            completion_pct = round((completed / max(total, 1)) * 100, 2)
            
            return {
                "total": total,
                "completed": completed,
                "completion_pct": completion_pct
            }
            
        except Exception as e:
            print(f"Error getting category completion for {client}: {e}")
            return {"total": 0, "completed": 0, "completion_pct": 0}
    
    def purge_results_data(
        self,
        older_than_days: Optional[int] = None,
        include_products: bool = False,
        delete_all: bool = False
    ) -> Dict[str, int]:
        """Remove historical results data with optional age filter."""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()

            stats = {
                "price_history_deleted": 0,
                "change_logs_deleted": 0,
                "products_deleted": 0
            }

            cutoff_iso = None
            if older_than_days is not None and not delete_all:
                cutoff = datetime.datetime.utcnow() - datetime.timedelta(days=older_than_days)
                cutoff_iso = cutoff.isoformat()

            if delete_all or cutoff_iso is None:
                cursor.execute('DELETE FROM ms_product_changes')
                stats["change_logs_deleted"] = cursor.rowcount

                cursor.execute('DELETE FROM ms_price_history')
                stats["price_history_deleted"] = cursor.rowcount

                if include_products:
                    cursor.execute('DELETE FROM ms_products')
                    stats["products_deleted"] = cursor.rowcount
            else:
                cursor.execute(
                    'DELETE FROM ms_product_changes WHERE changed_at < ?',
                    (cutoff_iso,)
                )
                stats["change_logs_deleted"] = cursor.rowcount

                cursor.execute(
                    'DELETE FROM ms_price_history WHERE recorded_at < ?',
                    (cutoff_iso,)
                )
                stats["price_history_deleted"] = cursor.rowcount

                if include_products:
                    cursor.execute(
                        'DELETE FROM ms_products WHERE updated_at < ?',
                        (cutoff_iso,)
                    )
                    stats["products_deleted"] = cursor.rowcount

            conn.commit()
            return stats

        except Exception as e:
            conn.rollback()
            print(f"Error purging results data: {e}")
            return {
                "price_history_deleted": 0,
                "change_logs_deleted": 0,
                "products_deleted": 0
            }

    def get_last_run(self, client: str) -> Optional[str]:
        """Get last scraping run timestamp for client"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            if client == "mobilesentrix":
                cursor.execute('''
                    SELECT started_at FROM scraper_runs 
                    WHERE status = 'completed'
                    ORDER BY started_at DESC LIMIT 1
                ''')
                result = cursor.fetchone()
                return result["started_at"] if result else None
            elif client == "xcellparts":
                cursor.execute('''
                    SELECT MAX(timestamp) as last_run FROM fetch_history
                    WHERE json_extract(rules, '$.site') = 'xcellparts'
                ''')
                result = cursor.fetchone()
                return result["last_run"] if result else None
            elif client == "txparts":
                cursor.execute('''
                    SELECT MAX(timestamp) as last_run FROM fetch_history
                    WHERE json_extract(rules, '$.site') = 'txparts'
                ''')
                result = cursor.fetchone()
                return result["last_run"] if result else None
            else:
                return None
                
        except Exception as e:
            print(f"Error getting last run for {client}: {e}")
            return None
    
    def get_next_run_eta_minutes(self, client: str, cadence_hours: int = 12) -> int:
        """Calculate ETA for next run in minutes"""
        try:
            last_run = self.get_last_run(client)
            if not last_run:
                return 0  # Run immediately if never run

            import datetime

            # Normalize last run into a timezone-aware datetime so math never mixes aware/naive
            if isinstance(last_run, datetime.datetime):
                last_dt = last_run
            else:
                # Handle strings that may be in ISO format with or without a trailing Z
                last_str = str(last_run).strip()
                if last_str.endswith('Z'):
                    last_str = last_str[:-1] + '+00:00'
                last_dt = datetime.datetime.fromisoformat(last_str)

            if last_dt.tzinfo is None:
                last_dt = last_dt.replace(tzinfo=datetime.timezone.utc)

            next_dt = last_dt + datetime.timedelta(hours=cadence_hours)
            now_dt = datetime.datetime.now(datetime.timezone.utc)

            eta_minutes = int((next_dt - now_dt).total_seconds() / 60)
            return max(0, eta_minutes)  # Never negative

        except Exception as e:
            print(f"Error calculating next run ETA for {client}: {e}")
            return 0
    
    def get_recent_changes(self, client: str, limit: int = 50, offset: int = 0,
                          change_types: Optional[List[str]] = None, from_date: str = None,
                          to_date: str = None, search_query: str = None, with_total: bool = False) -> Any:
        """Get recent changes with filters"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()

            change_filters = [ct.lower() for ct in change_types] if change_types else []
            search_term = (search_query or '').strip().lower()

            if client == "mobilesentrix":
                normalization_case = self._build_change_type_case_sql('pc.change_type')
                significant_predicate = self._build_significant_change_predicate('pc')
                baseline_exclusion = self._build_baseline_exclusion_predicate('pc', 'p')

                price_subquery = '''
                    SELECT
                        ph.id AS change_id,
                        ph.product_id AS product_id,
                        'price' AS change_type,
                        ph.recorded_at AS changed_at,
                        prev.price AS old_numeric,
                        ph.price AS new_numeric,
                        NULL AS old_text,
                        NULL AS new_text,
                        NULL AS metadata,
                        p.sku AS sku,
                        p.title AS title,
                        p.product_url AS url,
                        b.name AS brand_name,
                        cat.name AS category_name,
                        m.name AS model_name,
                        p.compatibility AS compatibility,
                        p.variant_details AS variant_details,
                        p.stock_status AS stock_status,
                        p.description AS description
                    FROM ms_price_history ph
                    JOIN ms_products p ON ph.product_id = p.id
                    JOIN ms_models m ON p.model_id = m.id
                    JOIN ms_categories cat ON m.category_id = cat.id
                    JOIN ms_brands b ON cat.brand_id = b.id
                    LEFT JOIN ms_price_history prev
                        ON prev.product_id = ph.product_id
                       AND prev.recorded_at = (
                            SELECT MAX(recorded_at)
                            FROM ms_price_history
                            WHERE product_id = ph.product_id
                              AND recorded_at < ph.recorded_at
                        )
                    WHERE prev.recorded_at IS NOT NULL
                      AND (
                          (prev.price IS NULL AND ph.price IS NOT NULL)
                          OR (prev.price IS NOT NULL AND ph.price IS NULL)
                          OR (
                              prev.price IS NOT NULL
                              AND ph.price IS NOT NULL
                              AND prev.price != ph.price
                          )
                      )
                '''

                property_subquery = f'''
                    SELECT *
                    FROM (
                        SELECT
                            pc.id AS change_id,
                            pc.product_id AS product_id,
                            {normalization_case} AS change_type,
                            pc.changed_at AS changed_at,
                            NULL AS old_numeric,
                            NULL AS new_numeric,
                            pc.old_value AS old_text,
                            pc.new_value AS new_text,
                            pc.metadata AS metadata,
                            p.sku AS sku,
                            p.title AS title,
                            p.product_url AS url,
                            b.name AS brand_name,
                            cat.name AS category_name,
                            m.name AS model_name,
                            p.compatibility AS compatibility,
                            p.variant_details AS variant_details,
                            p.stock_status AS stock_status,
                            p.description AS description
                        FROM ms_product_changes pc
                        JOIN ms_products p ON pc.product_id = p.id
                        JOIN ms_models m ON p.model_id = m.id
                        JOIN ms_categories cat ON m.category_id = cat.id
                        JOIN ms_brands b ON cat.brand_id = b.id
                        WHERE {significant_predicate}
                          AND NOT {baseline_exclusion}
                    ) typed_changes
                    WHERE typed_changes.change_type != 'price'
                '''

                union_sql = f"{price_subquery}\nUNION ALL\n{property_subquery}"

                where_clauses: List[str] = []
                params: List[Any] = []

                if change_filters:
                    placeholders = ','.join('?' for _ in change_filters)
                    where_clauses.append(f"LOWER(combined.change_type) IN ({placeholders})")
                    params.extend(change_filters)

                if from_date:
                    where_clauses.append("combined.changed_at >= ?")
                    params.append(from_date)

                if to_date:
                    where_clauses.append("combined.changed_at <= ?")
                    params.append(to_date)

                if search_term:
                    where_clauses.append("(LOWER(combined.title) LIKE ? OR LOWER(combined.sku) LIKE ?)")
                    like_term = f"%{search_term}%"
                    params.extend([like_term, like_term])

                where_sql = ''
                if where_clauses:
                    where_sql = 'WHERE ' + ' AND '.join(where_clauses)

                count_query = f"SELECT COUNT(*) AS total FROM ({union_sql}) AS combined {where_sql}"
                cursor.execute(count_query, params)
                count_row = cursor.fetchone()
                total = count_row['total'] if count_row else 0

                data_query = f'''
                    SELECT *
                    FROM ({union_sql}) AS combined
                    {where_sql}
                    ORDER BY combined.changed_at DESC
                    LIMIT ? OFFSET ?
                '''
                data_params = params + [limit, offset]
                cursor.execute(data_query, data_params)
                rows = cursor.fetchall()

                changes: List[Dict[str, Any]] = []
                for row in rows:
                    compatibility = self._safe_load_json(row['compatibility'], [])
                    variant_details = self._safe_load_json(row['variant_details'], {})
                    metadata = self._safe_load_json(row['metadata'], {})

                    raw_old = row['old_numeric'] if row['old_numeric'] is not None else row['old_text']
                    raw_new = row['new_numeric'] if row['new_numeric'] is not None else row['new_text']

                    old_display = self._format_change_value(row['change_type'], row['old_numeric'], row['old_text'])
                    new_display = self._format_change_value(row['change_type'], row['new_numeric'], row['new_text'])

                    difference = None
                    direction = 'flat'
                    if row['change_type'] == 'price':
                        if row['old_numeric'] is not None and row['new_numeric'] is not None:
                            difference = row['new_numeric'] - row['old_numeric']
                        elif row['old_numeric'] is not None and row['new_numeric'] is None:
                            difference = -row['old_numeric']
                        elif row['old_numeric'] is None and row['new_numeric'] is not None:
                            difference = row['new_numeric']

                        if isinstance(difference, (int, float)):
                            if difference > 0:
                                direction = 'up'
                            elif difference < 0:
                                direction = 'down'
                        else:
                            difference = None

                    product_title = self._sanitize_product_title(row['title'])
                    sku = (row['sku'] or 'N/A').strip() or 'N/A'
                    category_name = self._clean_category_name(row['category_name'])
                    model_identifier = self._build_model_identifier(row['brand_name'], row['model_name'], compatibility)
                    chipset = self._extract_chipset(variant_details, metadata, row['title'])

                    delta_display = self._format_change_delta_text(row['change_type'], difference)
                    change_label = self._build_change_label(row['change_type'], old_display, new_display, difference)

                    changes.append({
                        "product_title": product_title,
                        "title": product_title,
                        "sku": sku,
                        "url": row['url'],
                        "product_url": row['url'],
                        "change_type": row['change_type'],
                        "change_label": change_label,
                        "old_value": old_display,
                        "new_value": new_display,
                        "old_value_raw": raw_old,
                        "new_value_raw": raw_new,
                        "old_numeric": row['old_numeric'],
                        "new_numeric": row['new_numeric'],
                        "difference": difference,
                        "difference_direction": direction,
                        "delta_display": delta_display,
                        "changed_at": row['changed_at'],
                        "timestamp": row['changed_at'],
                        "brand": row['brand_name'],
                        "category": category_name,
                        "model": row['model_name'],
                        "model_identifier": model_identifier,
                        "chipset": chipset,
                        "compatibility": compatibility,
                        "variant_details": variant_details,
                        "metadata": metadata,
                        "stock_status": row['stock_status'],
                        "description_snapshot": row['description']
                    })

                if with_total:
                    result_payload: Any = {"items": changes, "total": total}
                else:
                    result_payload = changes

                has_no_changes = (with_total and total == 0) or (not with_total and not changes)
                if has_no_changes and client == "mobilesentrix":
                    baseline_payload = self._get_recent_baseline_snapshots(limit, offset, with_total)
                    return baseline_payload

                return result_payload

            # Fallback for other clients (legacy behaviour)
            site_pattern = '%xcell%'
            if client == 'txparts':
                site_pattern = '%txparts%'

            sql = '''
                SELECT 
                    url,
                    title,
                    site,
                    price_text as new_value,
                    created_at as changed_at
                FROM items
                WHERE site LIKE ?
            '''
            params = [site_pattern]

            if from_date:
                sql += ' AND created_at >= ?'
                params.append(from_date)

            if to_date:
                sql += ' AND created_at <= ?'
                params.append(to_date)

            if search_query:
                sql += ' AND title LIKE ?'
                params.append(f'%{search_query}%')

            sql += ' ORDER BY created_at DESC LIMIT ? OFFSET ?'
            params.extend([limit, offset])

            cursor.execute(sql, params)
            rows = cursor.fetchall()

            legacy_changes = []
            for row in rows:
                product_title = self._sanitize_product_title(row['title'])
                legacy_changes.append({
                    "sku": "N/A",
                    "product_title": product_title,
                    "title": product_title,
                    "url": row['url'],
                    "product_url": row['url'],
                    "change_type": "price",
                    "change_label": "Price changed",
                    "old_value": "N/A",
                    "new_value": row['new_value'],
                    "old_value_raw": None,
                    "new_value_raw": row['new_value'],
                    "old_numeric": None,
                    "new_numeric": None,
                    "difference": None,
                    "difference_direction": "flat",
                    "delta_display": None,
                    "changed_at": row['changed_at'],
                    "timestamp": row['changed_at'],
                    "category": '',
                    "brand": row['site'] if 'site' in row.keys() else '',
                    "model": '',
                    "compatibility": [],
                    "variant_details": {},
                    "metadata": {},
                    "stock_status": '',
                    "description_snapshot": None
                })

            if with_total:
                count_sql = 'SELECT COUNT(*) as total FROM items WHERE site LIKE ?'
                cursor.execute(count_sql, [site_pattern])
                total_row = cursor.fetchone()
                total = total_row['total'] if total_row else len(legacy_changes)
                return {"items": legacy_changes, "total": total}

            return legacy_changes

        except Exception as e:
            print(f"Error getting recent changes for {client}: {e}")
            return {"items": [], "total": 0} if with_total else []

    def _get_recent_baseline_snapshots(self, limit: int, offset: int, with_total: bool) -> Any:
        """Return baseline product snapshots so the dashboard shows data before live diffs exist."""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()

            cursor.execute('SELECT COUNT(*) AS total FROM ms_products')
            total = cursor.fetchone()["total"] or 0

            cursor.execute('''
                SELECT
                    p.id AS product_id,
                    p.sku,
                    p.title,
                    p.product_url,
                    p.price,
                    p.stock_status,
                    p.description,
                    p.updated_at,
                    p.created_at,
                    p.compatibility,
                    p.variant_details,
                    b.name AS brand_name,
                    c.name AS category_name,
                    m.name AS model_name,
                    baseline.captured_at AS captured_at
                FROM ms_products p
                JOIN ms_models m ON p.model_id = m.id
                JOIN ms_categories c ON m.category_id = c.id
                JOIN ms_brands b ON c.brand_id = b.id
                LEFT JOIN ms_product_baselines baseline ON baseline.product_id = p.id
                ORDER BY datetime(p.updated_at) DESC, datetime(p.created_at) DESC
                LIMIT ? OFFSET ?
            ''', (limit, offset))

            rows = cursor.fetchall()
            items: List[Dict[str, Any]] = []

            for row in rows:
                price_value = row['price'] if row['price'] is not None else None
                price_display = self._format_change_value('price', price_value, None)
                stock_display = self._format_change_value('stock', None, row['stock_status']) if row['stock_status'] else '—'

                captured_at = row['captured_at'] or row['updated_at'] or row['created_at']
                product_title = self._sanitize_product_title(row['title'])
                sku = (row['sku'] or 'N/A').strip() or 'N/A'
                category_name = self._clean_category_name(row['category_name'])
                compatibility = self._safe_load_json(row['compatibility'], [])
                variant_details = self._safe_load_json(row['variant_details'], {})

                items.append({
                    "product_title": product_title,
                    "title": product_title,
                    "sku": sku,
                    "url": row['product_url'],
                    "product_url": row['product_url'],
                    "change_type": 'baseline',
                    "change_label": 'Initial snapshot captured',
                    "old_value": '—',
                    "new_value": price_display,
                    "old_value_raw": None,
                    "new_value_raw": price_value,
                    "old_numeric": None,
                    "new_numeric": price_value,
                    "difference": None,
                    "difference_direction": 'flat',
                    "delta_display": None,
                    "changed_at": captured_at,
                    "timestamp": captured_at,
                    "brand": row['brand_name'],
                    "category": category_name,
                    "model": row['model_name'],
                    "model_identifier": self._build_model_identifier(row['brand_name'], row['model_name'], compatibility),
                    "chipset": self._extract_chipset(variant_details, {}, row['title']),
                    "compatibility": compatibility,
                    "variant_details": variant_details,
                    "metadata": {
                        "stock_status": stock_display
                    },
                    "stock_status": row['stock_status'] or '',
                    "description_snapshot": row['description'] or '',
                    "is_baseline": True
                })

            return {"items": items, "total": total} if with_total else items

        except Exception as baseline_error:
            print(f"Error building baseline fallback: {baseline_error}")
            return {"items": [], "total": 0} if with_total else []
    
    def export_changes_to_xlsx(self, client: str, filters: Dict = None) -> io.BytesIO:
        """Export changes to XLSX format"""
        try:
            from openpyxl import Workbook
            
            # Get changes with filters
            changes = self.get_recent_changes(
                client=client,
                limit=10000,  # Export more data
                change_types=filters.get('change_types', []) if filters else [],
                from_date=filters.get('from_date') if filters else None,
                to_date=filters.get('to_date') if filters else None,
                search_query=filters.get('search_query') if filters else None
            )
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"{client.title()} Changes"
            
            # Headers
            headers = ["SKU", "Title", "URL", "Change Type", "Old Value", "New Value", "Changed At"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data rows
            for row_idx, change in enumerate(changes, 2):
                ws.cell(row=row_idx, column=1, value=change.get("sku", ""))
                ws.cell(row=row_idx, column=2, value=change.get("product_title") or change.get("title", ""))
                ws.cell(row=row_idx, column=3, value=change.get("product_url") or change.get("url", ""))
                ws.cell(row=row_idx, column=4, value=change.get("change_type", ""))
                ws.cell(row=row_idx, column=5, value=change.get("old_value", ""))
                ws.cell(row=row_idx, column=6, value=change.get("new_value", ""))
                ws.cell(row=row_idx, column=7, value=change.get("changed_at", ""))
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
            
        except Exception as e:
            print(f"Error exporting changes to XLSX for {client}: {e}")
            return io.BytesIO()

def _default_db_path(filename: str) -> str:
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, filename)


EXTRACTOR_DB_PATH = (
    os.environ.get("EXTRACTOR_DATABASE_PATH")
    or os.environ.get("DATABASE_PATH")
    or _default_db_path("mobilesentrix.db")
)

RESULTS_DB_PATH = (
    os.environ.get("RESULTS_DATABASE_PATH")
    or _default_db_path("mobilesentrix_results.db")
)


# Legacy extractor database manager (history + items)
db_manager = DatabaseManager(
    db_path=EXTRACTOR_DB_PATH,
    create_legacy_tables=True,
    create_results_tables=False
)


# Results dashboard database manager (ms_* tables plus xcell legacy tables)
results_db_manager = DatabaseManager(
    db_path=RESULTS_DB_PATH,
    create_legacy_tables=True,
    create_results_tables=True
)
